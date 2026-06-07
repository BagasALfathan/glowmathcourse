from django.contrib import messages
from django.core.paginator import Paginator
from django.db import models
from django.db.models import Count, Q
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from academics.models import AcademicPeriod, Category, Kelas, KelasStatus, Quarter, Schedule, Subject
from academics.utils import build_schedule_grid, _COLOR_PALETTE, _SCHEDULE_DAYS
from accounts.decorators import role_required
from accounts.models import (
    AdminProfile, ApprovalStatus, Level, Role, StudentProfile, TeacherProfile, User,
)
from enrollments.models import Enrollment, EnrollmentStatus
from grades.models import Grade, GradeType
from ratings.models import TeacherRating
from activity_logs.models import ActivityLog
from activity_logs.utils import log_activity

PAGE_SIZE = 20

_STATUS_TABLE_ID = {
    ApprovalStatus.PENDING: 'pending-table-wrapper',
    ApprovalStatus.APPROVED: 'approved-table-wrapper',
    ApprovalStatus.REJECTED: 'rejected-table-wrapper',
}


# ── User Approval ──────────────────────────────────────────────────────────────

def _user_wa_url(user):
    """Return a wa.me URL using the user's profile phone, or None."""
    try:
        if user.role == Role.STUDENT:
            phone = user.student_profile.phone
        elif user.role == Role.TEACHER:
            phone = user.teacher_profile.phone
        else:
            return None
    except Exception:
        return None
    if not phone:
        return None
    p = phone.strip().replace(' ', '').replace('-', '').replace('+', '')
    if p.startswith('0'):
        p = '62' + p[1:]
    return f'https://wa.me/{p}' if p else None


@role_required('ADMIN')
def pending_users_view(request):
    """Data Pro v5: approval cards. Filter by role + search, no tabs."""
    q = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '')

    qs = (
        User.objects
        .filter(approval_status=ApprovalStatus.PENDING, is_deleted=False)
        .select_related('student_profile', 'teacher_profile')
        .order_by('-date_joined')
    )
    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
            | Q(username__icontains=q)
        )
    if role_filter in [Role.STUDENT, Role.TEACHER]:
        qs = qs.filter(role=role_filter)

    pending_users = list(qs)
    for u in pending_users:
        u.wa_url = _user_wa_url(u)

    # Total pending across all roles (badge in sidebar already uses
    # the global context processor, but we also surface it here so
    # the page header counts match the unfiltered universe).
    pending_count = User.objects.filter(
        approval_status=ApprovalStatus.PENDING, is_deleted=False
    ).count()

    return render(request, 'admin_panel/pending_users.html', {
        'pending_users': pending_users,
        'pending_count': pending_count,
        'q': q,
        'role_filter': role_filter,
    })


@role_required('ADMIN')
def users_table_partial(request):
    status = request.GET.get('status', ApprovalStatus.PENDING)
    if status not in [ApprovalStatus.PENDING, ApprovalStatus.APPROVED, ApprovalStatus.REJECTED]:
        status = ApprovalStatus.PENDING

    q = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '')
    page_num = request.GET.get('page', 1)

    qs = (
        User.objects
        .filter(approval_status=status, is_deleted=False)
        .select_related('student_profile', 'teacher_profile')
    )
    if status == ApprovalStatus.APPROVED:
        qs = qs.filter(role__in=[Role.STUDENT, Role.TEACHER])

    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
            | Q(username__icontains=q)
        )

    if role_filter in [Role.STUDENT, Role.TEACHER]:
        qs = qs.filter(role=role_filter)

    qs = qs.order_by('-date_joined')
    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(page_num)

    for user in page_obj.object_list:
        user.wa_url = _user_wa_url(user)

    return render(request, 'admin_panel/_users_table.html', {
        'page_obj': page_obj,
        'status': status,
        'q': q,
        'role_filter': role_filter,
        'table_id': _STATUS_TABLE_ID.get(status, 'pending-table-wrapper'),
    })


@role_required('ADMIN')
@require_POST
def change_status_view(request, user_id):
    user = get_object_or_404(User, pk=user_id, is_deleted=False)
    new_status = request.POST.get('new_status')

    if new_status == ApprovalStatus.APPROVED:
        user.approval_status = ApprovalStatus.APPROVED
        user.is_active = True
        user.save()
        log_activity(request.user, 'approved', 'user', user.pk)
        messages.success(
            request,
            f'Akun {user.get_full_name() or user.username} berhasil disetujui.',
        )
    elif new_status == ApprovalStatus.REJECTED:
        user.approval_status = ApprovalStatus.REJECTED
        user.is_active = False
        user.save()
        log_activity(request.user, 'rejected', 'user', user.pk)
        messages.warning(
            request,
            f'Akun {user.get_full_name() or user.username} telah ditolak/dicabut.',
        )
    else:
        messages.error(request, 'Status tidak valid.')

    return redirect('admin_panel:pending_users')


# ── User Management ────────────────────────────────────────────────────────────

@role_required('ADMIN')
def users_list(request):
    """Full /admin-panel/users/ page. Reads the same q / role / status
    querystring as users_list_partial so deep links like
    /admin-panel/users/?q=gracia land already filtered on first paint.
    """
    valid_roles = {Role.STUDENT, Role.TEACHER, Role.ADMIN}
    valid_statuses = {'', 'active', 'inactive', 'rejected', 'deleted'}
    q = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '')
    if role_filter not in valid_roles:
        role_filter = ''
    status_filter = request.GET.get('status', '')
    if status_filter not in valid_statuses:
        status_filter = ''
    return render(request, 'admin_panel/users_list.html', {
        'q': q,
        'role_filter': role_filter,
        'status_filter': status_filter,
    })


@role_required('ADMIN')
def users_list_partial(request):
    q = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')
    page_num = request.GET.get('page', 1)

    if status_filter == 'deleted':
        qs = User.objects.filter(is_deleted=True).order_by('-deleted_at')
    elif status_filter == 'rejected':
        qs = User.objects.filter(
            is_deleted=False, approval_status=ApprovalStatus.REJECTED
        ).order_by('-date_joined')
    else:
        qs = User.objects.filter(is_deleted=False).order_by('-date_joined')

    if q:
        qs = qs.filter(
            Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(email__icontains=q)
            | Q(username__icontains=q)
        )

    if role_filter in [Role.STUDENT, Role.TEACHER, Role.ADMIN]:
        qs = qs.filter(role=role_filter)

    if status_filter == 'active':
        qs = qs.filter(is_active=True)
    elif status_filter == 'inactive':
        qs = qs.filter(is_active=False, is_deleted=False)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(page_num)

    return render(request, 'admin_panel/_users_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'role_filter': role_filter,
        'status_filter': status_filter,
    })


@role_required('ADMIN')
def user_detail(request, user_id):
    target_user = get_object_or_404(User, pk=user_id, is_deleted=False)

    enrollments = None
    classes_teaching = None

    if target_user.role == Role.STUDENT:
        from enrollments.models import Enrollment
        enrollments = (
            Enrollment.objects
            .filter(student_profile__user=target_user, is_deleted=False)
            .select_related('kelas__subject', 'kelas__teacher_profile__user')
            .order_by('-enrolled_at')
        )
    elif target_user.role == Role.TEACHER:
        from academics.models import Kelas
        classes_teaching = (
            Kelas.objects
            .filter(teacher_profile__user=target_user, is_deleted=False)
            .select_related('subject')
            .order_by('name')
        )

    return render(request, 'admin_panel/user_detail.html', {
        'target_user': target_user,
        'enrollments': enrollments,
        'classes_teaching': classes_teaching,
    })


@role_required('ADMIN')
def user_reset_password(request, user_id):
    target_user = get_object_or_404(User, pk=user_id, is_deleted=False)

    if request.method == 'POST':
        new_pw = request.POST.get('new_password', '').strip()
        confirm = request.POST.get('confirm_password', '').strip()

        if len(new_pw) < 8:
            messages.error(request, 'Kata sandi baru minimal 8 karakter.')
        elif new_pw != confirm:
            messages.error(request, 'Konfirmasi kata sandi tidak cocok.')
        else:
            target_user.set_password(new_pw)
            target_user.save()
            log_activity(request.user, 'updated', 'user', target_user.pk)
            messages.success(
                request,
                f'Password berhasil direset untuk {target_user.username}.',
            )
            return redirect('admin_panel:user_detail', user_id=user_id)

    return redirect('admin_panel:user_detail', user_id=user_id)


@role_required('ADMIN')
def user_edit(request, user_id):
    target_user = get_object_or_404(User, pk=user_id, is_deleted=False)

    profile = None
    if target_user.role == Role.STUDENT:
        profile, _ = StudentProfile.objects.get_or_create(user=target_user)
    elif target_user.role == Role.TEACHER:
        profile, _ = TeacherProfile.objects.get_or_create(user=target_user)
    elif target_user.role == Role.ADMIN:
        profile, _ = AdminProfile.objects.get_or_create(user=target_user)

    if request.method == 'POST':
        target_user.first_name = request.POST.get('first_name', '').strip()
        target_user.last_name = request.POST.get('last_name', '').strip()
        new_email = request.POST.get('email', '').strip()
        target_user.is_active = request.POST.get('is_active') == '1'
        new_approval = request.POST.get('approval_status', '')
        if new_approval in [ApprovalStatus.PENDING, ApprovalStatus.APPROVED, ApprovalStatus.REJECTED]:
            target_user.approval_status = new_approval

        if new_email and new_email != target_user.email:
            if User.objects.filter(email__iexact=new_email, is_deleted=False).exclude(pk=target_user.pk).exists():
                messages.error(request, 'Email sudah digunakan akun lain.')
                return render(request, 'admin_panel/user_edit.html', {
                    'target_user': target_user,
                    'profile': profile,
                    'approval_choices': ApprovalStatus.choices,
                })
            target_user.email = new_email

        target_user.save()

        if profile:
            if target_user.role == Role.STUDENT:
                profile.level = request.POST.get('level', '').strip()
                profile.school_name = request.POST.get('school_name', '').strip()
                profile.phone = request.POST.get('phone', '').strip()
                profile.parent_name = request.POST.get('parent_name', '').strip()
                profile.parent_phone = request.POST.get('parent_phone', '').strip()
                profile.address = request.POST.get('address', '').strip()
            elif target_user.role == Role.TEACHER:
                profile.specialization = request.POST.get('specialization', '').strip()
                profile.phone = request.POST.get('phone', '').strip()
                profile.experience_years = int(request.POST.get('experience_years', 0) or 0)
                profile.bio = request.POST.get('bio', '').strip()
            elif target_user.role == Role.ADMIN:
                profile.phone = request.POST.get('phone', '').strip()
            profile.save()

        messages.success(request, 'Pengguna berhasil diperbarui!')
        return redirect('admin_panel:user_detail', user_id=target_user.pk)

    return render(request, 'admin_panel/user_edit.html', {
        'target_user': target_user,
        'profile': profile,
        'approval_choices': ApprovalStatus.choices,
    })


@role_required('ADMIN')
@require_POST
def user_delete(request, user_id):
    target_user = get_object_or_404(User, pk=user_id, is_deleted=False)

    if target_user.pk == request.user.pk:
        messages.error(request, 'Anda tidak dapat menghapus akun Anda sendiri.')
        return redirect('admin_panel:user_detail', user_id=target_user.pk)

    name = target_user.get_full_name() or target_user.username
    target_user.is_deleted = True
    target_user.deleted_at = timezone.now()
    target_user.is_active = False
    target_user.save()
    log_activity(request.user, 'deleted', 'user', target_user.pk)

    messages.success(request, f'Pengguna {name} berhasil dihapus.')
    return redirect('admin_panel:users_list')


@role_required('ADMIN')
@require_POST
def user_restore(request, user_id):
    target_user = get_object_or_404(User, pk=user_id, is_deleted=True)
    name = target_user.get_full_name() or target_user.username
    target_user.is_deleted = False
    target_user.deleted_at = None
    target_user.is_active = True
    target_user.approval_status = ApprovalStatus.APPROVED
    target_user.save()
    log_activity(request.user, 'updated', 'user', target_user.pk)
    messages.success(request, f'Pengguna {name} berhasil dipulihkan.')
    return redirect('admin_panel:users_list')


@role_required('ADMIN')
@require_POST
def bulk_action(request):
    action = request.POST.get('bulk_action', '')
    user_ids = request.POST.getlist('selected_users')

    if not user_ids or action not in ['approve', 'reject']:
        messages.error(request, 'Pilih tindakan dan setidaknya satu pengguna.')
        return redirect('admin_panel:pending_users')

    users = User.objects.filter(pk__in=user_ids, is_deleted=False)
    count = users.count()

    if action == 'approve':
        users.update(approval_status=ApprovalStatus.APPROVED, is_active=True)
        messages.success(request, f'{count} akun berhasil disetujui.')
    else:
        users.update(approval_status=ApprovalStatus.REJECTED, is_active=False)
        messages.warning(request, f'{count} akun berhasil ditolak.')

    return redirect('admin_panel:pending_users')


# ── Categories ────────────────────────────────────────────────────────────────

@role_required('ADMIN')
def categories_list(request):
    return render(request, 'admin_panel/categories_list.html', {
        'q': request.GET.get('q', '').strip(),
        'status_filter': request.GET.get('status', ''),
    })


@role_required('ADMIN')
def categories_list_partial(request):
    """HTMX partial — search by name, filter by status, paginate."""
    q = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')

    qs = Category.objects.annotate(
        subject_count=models.Count('subjects', distinct=True),
    ).order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q))
    if status_filter == 'active':
        qs = qs.filter(is_active=True)
    elif status_filter == 'inactive':
        qs = qs.filter(is_active=False)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_categories_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'status_filter': status_filter,
    })


@role_required('ADMIN')
def category_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == '1'
        if not name:
            messages.error(request, 'Nama kategori wajib diisi.')
            return render(request, 'admin_panel/category_form.html', {
                'mode': 'create', 'category': None, 'form_data': request.POST,
            })
        Category.objects.create(name=name, description=description, is_active=is_active)
        messages.success(request, f'Kategori "{name}" berhasil ditambahkan.')
        return redirect('admin_panel:categories_list')
    return render(request, 'admin_panel/category_form.html', {
        'mode': 'create', 'category': None, 'form_data': None,
    })


@role_required('ADMIN')
def category_edit(request, category_id):
    category = get_object_or_404(Category, pk=category_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        is_active = request.POST.get('is_active') == '1'
        if not name:
            messages.error(request, 'Nama kategori wajib diisi.')
            return render(request, 'admin_panel/category_form.html', {
                'mode': 'edit', 'category': category, 'form_data': request.POST,
            })
        category.name = name
        category.description = description
        category.is_active = is_active
        category.save()
        messages.success(request, f'Kategori "{name}" berhasil diperbarui.')
        return redirect('admin_panel:categories_list')
    return render(request, 'admin_panel/category_form.html', {
        'mode': 'edit', 'category': category, 'form_data': None,
    })


@role_required('ADMIN')
@require_POST
def category_delete(request, category_id):
    category = get_object_or_404(Category, pk=category_id)
    if category.subjects.exists():
        messages.error(request, f'Kategori "{category.name}" tidak dapat dihapus karena memiliki mata pelajaran.')
        return redirect('admin_panel:categories_list')
    name = category.name
    category.delete()
    messages.success(request, f'Kategori "{name}" berhasil dihapus.')
    return redirect('admin_panel:categories_list')


# ── Subjects ──────────────────────────────────────────────────────────────────

@role_required('ADMIN')
def subjects_list(request):
    return render(request, 'admin_panel/subjects_list.html', {
        'q': request.GET.get('q', '').strip(),
        'category_filter': request.GET.get('category', ''),
        'status_filter': request.GET.get('status', ''),
        'categories': Category.objects.filter(is_active=True).order_by('name'),
    })


@role_required('ADMIN')
def subjects_list_partial(request):
    """HTMX partial — search by name, filter by category + status."""
    q = request.GET.get('q', '').strip()
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')

    qs = (
        Subject.objects
        .select_related('category')
        .annotate(class_count=models.Count(
            'classes', filter=Q(classes__is_deleted=False), distinct=True,
        ))
        .order_by('category__name', 'name')
    )
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(description__icontains=q))
    if category_filter.isdigit():
        qs = qs.filter(category_id=int(category_filter))
    if status_filter == 'active':
        qs = qs.filter(is_active=True)
    elif status_filter == 'inactive':
        qs = qs.filter(is_active=False)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_subjects_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'category_filter': category_filter,
        'status_filter': status_filter,
    })


def _subject_form_context(mode, subject=None, form_data=None):
    return {
        'mode': mode,
        'subject': subject,        # may be None
        'form_data': form_data,    # may be None
        'categories': Category.objects.filter(is_active=True).order_by('name'),
    }


@role_required('ADMIN')
def subject_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', '').strip()[:10]
        description = request.POST.get('description', '').strip()
        category_id = request.POST.get('category', '')
        is_active = request.POST.get('is_active') == '1'
        if not name or not category_id:
            messages.error(request, 'Nama dan kategori wajib diisi.')
            return render(request, 'admin_panel/subject_form.html',
                          _subject_form_context('create', form_data=request.POST))
        category = get_object_or_404(Category, pk=category_id)
        Subject.objects.create(
            name=name, icon=icon, description=description,
            category=category, is_active=is_active,
        )
        messages.success(request, f'Mata pelajaran "{name}" berhasil ditambahkan.')
        return redirect('admin_panel:subjects_list')
    return render(request, 'admin_panel/subject_form.html',
                  _subject_form_context('create'))


@role_required('ADMIN')
def subject_edit(request, subject_id):
    subject = get_object_or_404(Subject, pk=subject_id)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        icon = request.POST.get('icon', '').strip()[:10]
        description = request.POST.get('description', '').strip()
        category_id = request.POST.get('category', '')
        is_active = request.POST.get('is_active') == '1'
        if not name or not category_id:
            messages.error(request, 'Nama dan kategori wajib diisi.')
            return render(request, 'admin_panel/subject_form.html',
                          _subject_form_context('edit', subject=subject, form_data=request.POST))
        subject.name = name
        subject.icon = icon
        subject.description = description
        subject.category = get_object_or_404(Category, pk=category_id)
        subject.is_active = is_active
        subject.save()
        messages.success(request, f'Mata pelajaran "{name}" berhasil diperbarui.')
        return redirect('admin_panel:subjects_list')
    return render(request, 'admin_panel/subject_form.html',
                  _subject_form_context('edit', subject=subject))


@role_required('ADMIN')
@require_POST
def subject_delete(request, subject_id):
    subject = get_object_or_404(Subject, pk=subject_id)
    if subject.classes.filter(is_deleted=False).exists():
        messages.error(request, f'Mata pelajaran "{subject.name}" tidak dapat dihapus karena memiliki kelas aktif.')
        return redirect('admin_panel:subjects_list')
    name = subject.name
    subject.delete()
    messages.success(request, f'Mata pelajaran "{name}" berhasil dihapus.')
    return redirect('admin_panel:subjects_list')


# ── Academic Periods ──────────────────────────────────────────────────────────

@role_required('ADMIN')
def periods_list(request):
    return render(request, 'admin_panel/periods_list.html', {
        'q': request.GET.get('q', '').strip(),
        'year_filter': request.GET.get('year', ''),
        'type_filter': request.GET.get('type', ''),
        'all_years': list(AcademicPeriod.objects.order_by('-year')
                          .values_list('year', flat=True).distinct()),
    })


@role_required('ADMIN')
def periods_list_partial(request):
    from academics.models import PeriodType
    q = request.GET.get('q', '').strip()
    year_filter = request.GET.get('year', '')
    type_filter = request.GET.get('type', '')

    qs = AcademicPeriod.objects.order_by('-is_active', '-year', 'quarter', 'semester')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(year__icontains=q))
    if year_filter:
        qs = qs.filter(year=year_filter)
    if type_filter in [PeriodType.QUARTER, PeriodType.SEMESTER]:
        qs = qs.filter(period_type=type_filter)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_periods_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'year_filter': year_filter,
        'type_filter': type_filter,
    })


def _period_form_context(mode, period=None, form_data=None):
    from academics.models import PeriodType, Semester
    return {
        'mode': mode,
        'period': period,          # may be None
        'form_data': form_data,    # may be None
        'period_type_choices': PeriodType.choices,
        'quarter_choices': Quarter.choices,
        'semester_choices': Semester.choices,
    }


def _parse_period_post(post):
    from academics.models import PeriodType
    period_type = post.get('period_type', '').strip()
    data = {
        'year': post.get('year', '').strip(),
        'period_type': period_type,
        'quarter': post.get('quarter', '').strip() if period_type == PeriodType.QUARTER else '',
        'semester': post.get('semester', '').strip() if period_type == PeriodType.SEMESTER else '',
        'name': post.get('name', '').strip(),
        'start_date': post.get('start_date', '').strip(),
        'end_date': post.get('end_date', '').strip(),
        'is_active': post.get('is_active') == '1',
    }
    errors = []
    if not data['year']:
        errors.append('Tahun ajaran wajib diisi (contoh: 2026-2027).')
    if data['period_type'] not in [PeriodType.QUARTER, PeriodType.SEMESTER]:
        errors.append('Tipe periode tidak valid.')
    elif data['period_type'] == PeriodType.QUARTER and not data['quarter']:
        errors.append('Pilih Kuartal (Q1-Q4) untuk tipe Kuartal.')
    elif data['period_type'] == PeriodType.SEMESTER and not data['semester']:
        errors.append('Pilih Semester (Ganjil/Genap) untuk tipe Semester.')
    if not data['name']:
        errors.append('Nama periode wajib diisi.')
    if not data['start_date'] or not data['end_date']:
        errors.append('Tanggal mulai dan selesai wajib diisi.')
    return data, errors


@role_required('ADMIN')
def period_create(request):
    if request.method == 'POST':
        data, errors = _parse_period_post(request.POST)
        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'admin_panel/period_form.html',
                          _period_form_context('create', form_data=request.POST))
        dup_q = AcademicPeriod.objects.filter(
            year=data['year'], period_type=data['period_type'],
            quarter=data['quarter'], semester=data['semester'],
        )
        if dup_q.exists():
            messages.error(request, 'Kombinasi tahun + tipe + sub-periode sudah ada.')
            return render(request, 'admin_panel/period_form.html',
                          _period_form_context('create', form_data=request.POST))
        from django.db import transaction
        with transaction.atomic():
            if data['is_active']:
                AcademicPeriod.objects.filter(is_active=True).update(is_active=False)
            AcademicPeriod.objects.create(**data)
        messages.success(request, f'Periode "{data["name"]}" berhasil ditambahkan.')
        return redirect('admin_panel:periods_list')
    return render(request, 'admin_panel/period_form.html',
                  _period_form_context('create'))


@role_required('ADMIN')
def period_edit(request, period_id):
    period = get_object_or_404(AcademicPeriod, pk=period_id)
    if request.method == 'POST':
        data, errors = _parse_period_post(request.POST)
        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'admin_panel/period_form.html',
                          _period_form_context('edit', period=period, form_data=request.POST))
        dup_q = AcademicPeriod.objects.filter(
            year=data['year'], period_type=data['period_type'],
            quarter=data['quarter'], semester=data['semester'],
        ).exclude(pk=period_id)
        if dup_q.exists():
            messages.error(request, 'Kombinasi tahun + tipe + sub-periode sudah digunakan periode lain.')
            return render(request, 'admin_panel/period_form.html',
                          _period_form_context('edit', period=period, form_data=request.POST))
        from django.db import transaction
        with transaction.atomic():
            if data['is_active'] and not period.is_active:
                AcademicPeriod.objects.filter(is_active=True).exclude(pk=period_id).update(is_active=False)
            for k, v in data.items():
                setattr(period, k, v)
            period.save()
        messages.success(request, f'Periode "{data["name"]}" berhasil diperbarui.')
        return redirect('admin_panel:periods_list')
    return render(request, 'admin_panel/period_form.html',
                  _period_form_context('edit', period=period))


@role_required('ADMIN')
@require_POST
def period_set_active(request, period_id):
    from django.db import transaction
    period = get_object_or_404(AcademicPeriod, pk=period_id)
    with transaction.atomic():
        AcademicPeriod.objects.filter(is_active=True).exclude(pk=period_id).update(is_active=False)
        period.is_active = True
        period.save(update_fields=['is_active', 'updated_at'])
    messages.success(request, f'Periode "{period.name}" sekarang aktif.')
    return redirect('admin_panel:periods_list')


@role_required('ADMIN')
@require_POST
def period_delete(request, period_id):
    period = get_object_or_404(AcademicPeriod, pk=period_id)
    if period.classes.filter(is_deleted=False).exists():
        messages.error(request, f'Periode "{period.name}" tidak dapat dihapus karena memiliki kelas.')
        return redirect('admin_panel:periods_list')
    name = period.name
    period.delete()
    messages.success(request, f'Periode "{name}" berhasil dihapus.')
    return redirect('admin_panel:periods_list')


# ── Admin Classes ─────────────────────────────────────────────────────────────

@role_required('ADMIN')
def classes_list(request):
    periods = AcademicPeriod.objects.order_by('-year', 'quarter')
    return render(request, 'admin_panel/classes_list.html', {
        'periods': periods,
        'q': request.GET.get('q', '').strip(),
        'level_filter': request.GET.get('level', ''),
        'status_filter': request.GET.get('status', ''),
        'period_filter': request.GET.get('period', ''),
        'show_deleted': request.GET.get('show_deleted', '') == '1',
    })


@role_required('ADMIN')
def classes_list_partial(request):
    """Data Pro v5 — search by name / teacher, filter by level + status +
    period, explicit show_deleted toggle. Annotates enrolled count so the
    table can render "12/15" without N+1 queries.
    """
    q = request.GET.get('q', '').strip()
    level_filter = request.GET.get('level', '')
    status_filter = request.GET.get('status', '')
    period_filter = request.GET.get('period', '')
    show_deleted = request.GET.get('show_deleted', '') == '1'

    qs = (
        Kelas.objects
        .select_related('subject', 'teacher_profile__user', 'academic_period')
        .annotate(active_count=Count(
            'enrollments',
            filter=Q(enrollments__status=EnrollmentStatus.ACTIVE,
                     enrollments__is_deleted=False),
            distinct=True,
        ))
        .order_by('-created_at')
    )

    if show_deleted:
        qs = qs.filter(is_deleted=True)
    else:
        qs = qs.filter(is_deleted=False)

    if q:
        # Use the real FK path; PITFALLS.md flags @property shims like
        # Kelas.teacher as unusable in ORM filters.
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(teacher_profile__user__first_name__icontains=q)
            | Q(teacher_profile__user__last_name__icontains=q)
            | Q(teacher_profile__user__username__icontains=q)
            | Q(subject__name__icontains=q)
        )
    if level_filter in Level.values:
        qs = qs.filter(level=level_filter)
    if status_filter in ['OPEN', 'FULL', 'CLOSED']:
        qs = qs.filter(status=status_filter)
    if period_filter:
        qs = qs.filter(academic_period_id=period_filter)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_classes_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'level_filter': level_filter,
        'status_filter': status_filter,
        'period_filter': period_filter,
        'show_deleted': show_deleted,
    })


@role_required('ADMIN')
@require_POST
def class_change_status(request, kelas_id):
    kelas = get_object_or_404(Kelas, pk=kelas_id, is_deleted=False)
    new_status = request.POST.get('status', '')
    if new_status in ['OPEN', 'FULL', 'CLOSED']:
        kelas.status = new_status
        kelas.save()
    return redirect('admin_panel:classes_list')


@role_required('ADMIN')
@require_POST
def class_soft_delete(request, kelas_id):
    kelas = get_object_or_404(Kelas, pk=kelas_id, is_deleted=False)
    kelas.is_deleted = True
    kelas.deleted_at = timezone.now()
    kelas.status = 'CLOSED'
    kelas.save()
    log_activity(request.user, 'deleted', 'kelas', kelas.pk)
    messages.success(request, f'Kelas "{kelas.name}" berhasil dihapus.')
    return redirect('admin_panel:classes_list')


@role_required('ADMIN')
@require_POST
def class_restore(request, kelas_id):
    kelas = get_object_or_404(Kelas, pk=kelas_id, is_deleted=True)
    kelas.is_deleted = False
    kelas.deleted_at = None
    kelas.status = KelasStatus.CLOSED
    kelas.save()
    log_activity(request.user, 'updated', 'kelas', kelas.pk)
    messages.success(request, f'Kelas "{kelas.name}" berhasil dipulihkan.')
    return redirect('admin_panel:classes_list')


# ── Admin Enrollments ─────────────────────────────────────────────────────────

@role_required('ADMIN')
def enrollments_list(request):
    return render(request, 'admin_panel/enrollments_list.html', {
        'q': request.GET.get('q', '').strip(),
        'status_filter': request.GET.get('status', ''),
        'level_filter': request.GET.get('level', ''),
        'show_deleted': request.GET.get('show_deleted', '') == '1',
    })


@role_required('ADMIN')
def enrollments_list_partial(request):
    """Data Pro v5 — search siswa / kelas, filter status + level, explicit
    show_deleted toggle. select_related kelas__teacher_profile__user so the
    table can render the teacher column without N+1.
    """
    q = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '')
    level_filter = request.GET.get('level', '')
    show_deleted = request.GET.get('show_deleted', '') == '1'

    qs = (
        Enrollment.objects
        .select_related(
            'student_profile__user',
            'kelas__subject',
            'kelas__teacher_profile__user',
        )
        .order_by('-enrolled_at')
    )
    if show_deleted:
        qs = qs.filter(is_deleted=True)
    else:
        qs = qs.filter(is_deleted=False)

    if q:
        qs = qs.filter(
            Q(student_profile__user__first_name__icontains=q)
            | Q(student_profile__user__last_name__icontains=q)
            | Q(student_profile__user__username__icontains=q)
            | Q(kelas__name__icontains=q)
            | Q(kelas__teacher_profile__user__first_name__icontains=q)
            | Q(kelas__teacher_profile__user__last_name__icontains=q)
        )
    if status_filter in [EnrollmentStatus.ACTIVE, EnrollmentStatus.COMPLETED, EnrollmentStatus.DROPPED]:
        qs = qs.filter(status=status_filter)
    if level_filter in Level.values:
        qs = qs.filter(kelas__level=level_filter)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_enrollments_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'status_filter': status_filter,
        'level_filter': level_filter,
        'show_deleted': show_deleted,
    })


@role_required('ADMIN')
@require_POST
def enrollment_change_status(request, enrollment_id):
    enrollment = get_object_or_404(Enrollment, pk=enrollment_id, is_deleted=False)
    new_status = request.POST.get('status', '')
    if new_status in [EnrollmentStatus.ACTIVE, EnrollmentStatus.COMPLETED, EnrollmentStatus.DROPPED]:
        enrollment.status = new_status
        enrollment.save()
        log_activity(request.user, 'updated', 'enrollment', enrollment.pk)
    return redirect('admin_panel:enrollments_list')


@role_required('ADMIN')
@require_POST
def enrollment_soft_delete(request, enrollment_id):
    enrollment = get_object_or_404(Enrollment, pk=enrollment_id, is_deleted=False)
    enrollment.soft_delete()
    log_activity(request.user, 'deleted', 'enrollment', enrollment.pk)
    messages.success(request, 'Pendaftaran berhasil dihapus.')
    return redirect('admin_panel:enrollments_list')


@role_required('ADMIN')
@require_POST
def enrollment_restore(request, enrollment_id):
    enrollment = get_object_or_404(Enrollment, pk=enrollment_id, is_deleted=True)
    enrollment.is_deleted = False
    enrollment.deleted_at = None
    enrollment.save(update_fields=['is_deleted', 'deleted_at', 'updated_at'])
    log_activity(request.user, 'updated', 'enrollment', enrollment.pk)
    messages.success(request, 'Pendaftaran berhasil dipulihkan.')
    return redirect('admin_panel:enrollments_list')


@role_required('ADMIN')
@require_POST
def enrollment_bulk_action(request):
    """Atomic bulk action on enrollments (status change OR soft delete).

    Not to be confused with `bulk_action` (URL `enrollments/bulk-action/`),
    which is scoped to PENDING USERS approve/reject despite the misleading
    URL prefix. This endpoint is intentionally placed at a distinct path.
    """
    from django.db import transaction
    action = request.POST.get('bulk_action', '')
    ids_raw = request.POST.getlist('selected_ids')
    try:
        ids = [int(x) for x in ids_raw if str(x).strip().isdigit()]
    except (TypeError, ValueError):
        ids = []
    if not ids or not action:
        messages.error(request, 'Pilih tindakan dan setidaknya satu pendaftaran.')
        return redirect('admin_panel:enrollments_list')

    qs = Enrollment.objects.filter(pk__in=ids, is_deleted=False)
    count = qs.count()
    if count == 0:
        messages.error(request, 'Tidak ada pendaftaran valid yang dipilih.')
        return redirect('admin_panel:enrollments_list')

    with transaction.atomic():
        if action in [EnrollmentStatus.ACTIVE, EnrollmentStatus.COMPLETED, EnrollmentStatus.DROPPED]:
            qs.update(status=action)
            log_activity(request.user, 'updated', 'enrollment_bulk_status', None)
            messages.success(request, f'{count} pendaftaran diubah menjadi {action}.')
        elif action == 'soft_delete':
            qs.update(is_deleted=True, deleted_at=timezone.now())
            log_activity(request.user, 'deleted', 'enrollment_bulk', None)
            messages.success(request, f'{count} pendaftaran berhasil dihapus.')
        else:
            messages.error(request, 'Tindakan tidak valid.')

    return redirect('admin_panel:enrollments_list')


# ── Admin Grades ──────────────────────────────────────────────────────────────

@role_required('ADMIN')
def grades_list(request):
    return render(request, 'admin_panel/grades_list.html', {
        'grade_type_choices': GradeType.choices,
    })


@role_required('ADMIN')
def grades_list_partial(request):
    q = request.GET.get('q', '').strip()
    type_filter = request.GET.get('grade_type', '')
    page_num = request.GET.get('page', 1)

    qs = (
        Grade.objects
        .select_related(
            'enrollment__student_profile__user',
            'enrollment__kelas__subject',
        )
        .order_by('-graded_at')
    )

    if q:
        qs = qs.filter(
            Q(enrollment__student_profile__user__first_name__icontains=q)
            | Q(enrollment__student_profile__user__last_name__icontains=q)
            | Q(enrollment__kelas__name__icontains=q)
        )
    if type_filter in [GradeType.QUIZ, GradeType.MIDTERM, GradeType.FINAL, GradeType.ASSIGNMENT]:
        qs = qs.filter(grade_type=type_filter)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(page_num)

    return render(request, 'admin_panel/_grades_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'type_filter': type_filter,
    })


# ── Admin Ratings ─────────────────────────────────────────────────────────────

@role_required('ADMIN')
def ratings_list(request):
    return render(request, 'admin_panel/ratings_list.html', {
        'q': request.GET.get('q', '').strip(),
        'type_filter': request.GET.get('type', 'teacher'),
        'score_filter': request.GET.get('score', ''),
    })


@role_required('ADMIN')
def ratings_list_partial(request):
    """Data Pro v5 — type tab (teacher | class), search by comment / student /
    teacher / kelas name, score filter, paginated. Each row carries a delete
    form for moderation. Comments truncate in the cell; the modal-less design
    relies on title=full-text for hover preview."""
    from ratings.models import ClassRating
    q = request.GET.get('q', '').strip()
    type_filter = request.GET.get('type', 'teacher')
    if type_filter not in ('teacher', 'class'):
        type_filter = 'teacher'
    score_filter = request.GET.get('score', '')

    if type_filter == 'teacher':
        qs = (
            TeacherRating.objects
            .select_related(
                'enrollment__student_profile__user',
                'enrollment__kelas__subject',
                'teacher_profile__user',
            )
            .order_by('-created_at')
        )
        if q:
            qs = qs.filter(
                Q(comment__icontains=q)
                | Q(enrollment__student_profile__user__first_name__icontains=q)
                | Q(enrollment__student_profile__user__last_name__icontains=q)
                | Q(teacher_profile__user__first_name__icontains=q)
                | Q(teacher_profile__user__last_name__icontains=q)
            )
    else:
        qs = (
            ClassRating.objects
            .select_related(
                'enrollment__student_profile__user',
                'kelas__subject',
                'kelas__teacher_profile__user',
            )
            .order_by('-created_at')
        )
        if q:
            qs = qs.filter(
                Q(comment__icontains=q)
                | Q(enrollment__student_profile__user__first_name__icontains=q)
                | Q(enrollment__student_profile__user__last_name__icontains=q)
                | Q(kelas__name__icontains=q)
            )

    if score_filter in ['1', '2', '3', '4', '5']:
        qs = qs.filter(score=int(score_filter))

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_ratings_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'type_filter': type_filter,
        'score_filter': score_filter,
    })


@role_required('ADMIN')
@require_POST
def rating_delete(request, rating_type, rating_id):
    """Moderation: hard-delete a TeacherRating or ClassRating. Logged."""
    from ratings.models import ClassRating
    if rating_type == 'teacher':
        rating = get_object_or_404(TeacherRating, pk=rating_id)
        label = 'Penilaian guru'
    elif rating_type == 'class':
        rating = get_object_or_404(ClassRating, pk=rating_id)
        label = 'Penilaian kelas'
    else:
        messages.error(request, 'Tipe rating tidak valid.')
        return redirect('admin_panel:ratings_list')
    log_activity(request.user, 'deleted', f'{rating_type}_rating', rating.pk)
    rating.delete()
    messages.success(request, f'{label} berhasil dihapus.')
    return redirect('admin_panel:ratings_list')


# ── Admin Activity Logs ───────────────────────────────────────────────────────

LOGS_PAGE_SIZE = 30

LOG_TARGET_TYPES = [
    'kelas', 'enrollment', 'grade', 'attendance', 'rating',
    'teacher_rating', 'class_rating', 'user', 'announcement',
]

LOG_ACTIONS = [
    'created', 'updated', 'deleted', 'approved', 'rejected',
    'ADMIN_LOGIN', 'LOGIN', 'LOGOUT',
]

@role_required('ADMIN')
def logs_list(request):
    return render(request, 'admin_panel/logs_list.html', {
        'target_types': LOG_TARGET_TYPES,
        'actions': LOG_ACTIONS,
        'q': request.GET.get('q', '').strip(),
        'action_filter': request.GET.get('action', ''),
        'type_filter': request.GET.get('target_type', ''),
    })


@role_required('ADMIN')
def logs_list_partial(request):
    q = request.GET.get('q', '').strip()
    action_filter = request.GET.get('action', '')
    type_filter = request.GET.get('target_type', '')

    qs = ActivityLog.objects.select_related('user').order_by('-created_at')

    if q:
        qs = qs.filter(
            Q(user__first_name__icontains=q)
            | Q(user__last_name__icontains=q)
            | Q(user__username__icontains=q)
            | Q(action__icontains=q)
        )
    if action_filter in LOG_ACTIONS:
        qs = qs.filter(action=action_filter)
    if type_filter in LOG_TARGET_TYPES:
        qs = qs.filter(target_type=type_filter)

    paginator = Paginator(qs, LOGS_PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_logs_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'action_filter': action_filter,
        'type_filter': type_filter,
    })


# ── Admin Exports ──────────────────────────────────────────────────────────────

@role_required('ADMIN')
def export_students_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    students = (
        User.objects
        .filter(role=Role.STUDENT, is_deleted=False)
        .select_related('student_profile')
        .order_by('last_name', 'first_name')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Siswa'

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center')

    headers = [
        'No', 'Nama Lengkap', 'Email', 'Username', 'Jenjang',
        'Sekolah', 'Kelas Sekolah', 'No HP Siswa',
        'Nama Orang Tua', 'No HP Orang Tua', 'Status', 'Tanggal Daftar',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    from django.utils.timezone import localtime
    for i, user in enumerate(students, 1):
        try:
            p = user.student_profile
        except Exception:
            p = None
        status = 'Aktif' if user.is_active else 'Tidak Aktif'
        ws.append([
            i,
            user.get_full_name(),
            user.email,
            user.username,
            p.level if p else '-',
            p.school_name if p else '-',
            str(p.school_grade) if p and p.school_grade else '-',
            p.phone if p else '-',
            p.parent_name if p else '-',
            p.parent_phone if p else '-',
            status,
            localtime(user.date_joined).strftime('%d/%m/%Y'),
        ])

    col_widths = [5, 28, 28, 18, 8, 28, 12, 16, 24, 16, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from django.utils import timezone as tz
    filename = f'Data_Siswa_{tz.localdate().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@role_required('ADMIN')
def export_classes_excel(request):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from django.db.models import Count, Q

    from enrollments.models import EnrollmentStatus as ES

    klasses = (
        Kelas.objects
        .filter(is_deleted=False)
        .select_related('subject', 'teacher_profile__user', 'academic_period')
        .annotate(
            student_count=Count(
                'enrollments',
                filter=Q(enrollments__status=ES.ACTIVE, enrollments__is_deleted=False),
                distinct=True,
            )
        )
        .order_by('name')
    )

    wb = Workbook()
    ws = wb.active
    ws.title = 'Data Kelas'

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center')

    headers = [
        'No', 'Nama Kelas', 'Mata Pelajaran', 'Guru', 'Jenjang',
        'Kapasitas', 'Jumlah Siswa', 'Status', 'Periode',
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    status_map = {'OPEN': 'Buka', 'FULL': 'Penuh', 'CLOSED': 'Tutup'}
    for i, kelas in enumerate(klasses, 1):
        ws.append([
            i,
            kelas.name,
            kelas.subject.name if kelas.subject_id else '-',
            kelas.teacher.get_full_name() if kelas.teacher_id else '-',
            kelas.level,
            kelas.capacity,
            kelas.student_count,
            status_map.get(kelas.status, kelas.status),
            kelas.academic_period.name if kelas.academic_period_id else '-',
        ])

    col_widths = [5, 30, 22, 26, 8, 10, 14, 10, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from django.utils import timezone as tz
    filename = f'Data_Kelas_{tz.localdate().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ── Schedule views ─────────────────────────────────────────────────────────────

def _admin_schedule_ctx(request):
    teacher_filter = request.GET.get('teacher', '').strip()
    level_filter   = request.GET.get('level',   '').strip()
    room_filter    = request.GET.get('room',    '').strip()

    qs = (
        Schedule.objects
        .filter(
            kelas__is_deleted=False,
            kelas__status__in=[KelasStatus.OPEN, KelasStatus.FULL],
        )
        .select_related('kelas__subject__category', 'kelas__teacher_profile__user',
                        'kelas__academic_period')
        .order_by('kelas__teacher_profile__user__last_name', 'kelas__teacher_profile__user__first_name',
                  'start_time')
    )

    if teacher_filter:
        try:
            qs = qs.filter(kelas__teacher_profile__user_id=int(teacher_filter))
        except ValueError:
            teacher_filter = ''
    if level_filter:
        qs = qs.filter(kelas__level=level_filter)
    if room_filter:
        qs = qs.filter(room__icontains=room_filter)

    # Assign a stable color per teacher (ordered by pk)
    teacher_pks = list(
        User.objects.filter(role=Role.TEACHER, is_active=True, is_deleted=False)
        .order_by('pk').values_list('pk', flat=True)
    )
    teacher_color_map = {
        pk: _COLOR_PALETTE[idx % len(_COLOR_PALETTE)]
        for idx, pk in enumerate(teacher_pks)
    }

    items = []
    for sched in qs:
        kelas = sched.kelas
        color = teacher_color_map.get(kelas.teacher_profile.user_id, _COLOR_PALETTE[0])
        items.append({'schedule': sched, 'kelas': kelas, 'color': color})

    grid_rows, days_list = build_schedule_grid(items)

    teachers = list(
        User.objects.filter(role=Role.TEACHER, is_active=True, is_deleted=False)
        .order_by('first_name', 'last_name')
    )
    legend = [
        {'teacher': t, 'color': teacher_color_map.get(t.pk, _COLOR_PALETTE[0])}
        for t in teachers
        if any(i['kelas'].teacher_profile.user_id == t.pk for i in items)
    ]

    return {
        'grid_rows': grid_rows,
        'days_list': days_list,
        'days': _SCHEDULE_DAYS,
        'view_role': 'admin',
        'teachers': teachers,
        'legend': legend,
        'teacher_filter': teacher_filter,
        'level_filter': level_filter,
        'room_filter': room_filter,
        'total_slots': len(items),
    }


@role_required('ADMIN')
def admin_schedule(request):
    ctx = _admin_schedule_ctx(request)
    return render(request, 'admin_panel/admin_schedule.html', ctx)


@role_required('ADMIN')
def admin_schedule_print(request):
    ctx = _admin_schedule_ctx(request)
    return render(request, 'admin_panel/admin_schedule_print.html', ctx)


# ── Enrollment Progress (admin view) ──────────────────────────────────────────

@role_required('ADMIN')
def enrollment_progress(request, enrollment_id):
    from grades.views import _build_progress_ctx
    enrollment = get_object_or_404(
        Enrollment.objects.select_related(
            'student_profile__user', 'kelas__subject',
            'kelas__teacher_profile__user', 'kelas__academic_period',
        ),
        pk=enrollment_id,
    )
    ctx = _build_progress_ctx(enrollment)
    ctx['back_url'] = '/admin-panel/enrollments/'
    ctx['is_admin_view'] = True

    # Available target classes for transfer (same level, active, not this class)
    from django.db.models import Count as _Count, Q as _Q
    target_classes = (
        Kelas.objects
        .filter(
            level=enrollment.kelas.level,
            is_deleted=False,
            status__in=[KelasStatus.OPEN, KelasStatus.FULL],
        )
        .exclude(pk=enrollment.kelas.pk)
        .annotate(
            active_count=_Count(
                'enrollments',
                filter=_Q(enrollments__status=EnrollmentStatus.ACTIVE, enrollments__is_deleted=False),
            )
        )
        .filter(active_count__lt=models.F('capacity'))
        .select_related('subject')
        .order_by('name')
    )
    ctx['target_classes'] = target_classes
    ctx['enrollment'] = enrollment
    return render(request, 'admin_panel/enrollment_progress.html', ctx)


@role_required('ADMIN')
@require_POST
def enrollment_transfer(request, enrollment_id):
    enrollment = get_object_or_404(
        Enrollment, pk=enrollment_id, is_deleted=False, status=EnrollmentStatus.ACTIVE
    )
    target_kelas_id = request.POST.get('target_kelas', '')
    try:
        target_kelas = Kelas.objects.get(pk=int(target_kelas_id), is_deleted=False)
    except (ValueError, Kelas.DoesNotExist):
        messages.error(request, 'Kelas tujuan tidak valid.')
        return redirect('admin_panel:enrollment_progress', enrollment_id=enrollment_id)

    # Check level match
    try:
        student_level = enrollment.student.student_profile.level
    except Exception:
        student_level = ''
    if student_level and target_kelas.level != student_level:
        messages.error(request, 'Jenjang kelas tujuan tidak sesuai dengan jenjang siswa.')
        return redirect('admin_panel:enrollment_progress', enrollment_id=enrollment_id)

    # Check capacity
    active_count = Enrollment.objects.filter(
        kelas=target_kelas, status=EnrollmentStatus.ACTIVE, is_deleted=False
    ).count()
    if active_count >= target_kelas.capacity:
        messages.error(request, 'Kelas tujuan sudah penuh.')
        return redirect('admin_panel:enrollment_progress', enrollment_id=enrollment_id)

    # Check duplicate
    if Enrollment.objects.filter(
        student_profile__user=enrollment.student, kelas=target_kelas, is_deleted=False
    ).exists():
        messages.error(request, 'Siswa sudah terdaftar di kelas tujuan.')
        return redirect('admin_panel:enrollment_progress', enrollment_id=enrollment_id)

    student_name = enrollment.student.get_full_name() or enrollment.student.username
    old_class_name = enrollment.kelas.name

    # Drop current enrollment
    enrollment.status = EnrollmentStatus.DROPPED
    enrollment.save(update_fields=['status'])

    # Create new enrollment
    new_enrollment = Enrollment.objects.create(
        student_profile=enrollment.student_profile,
        kelas=target_kelas,
        status=EnrollmentStatus.ACTIVE,
        price_at_enrollment=target_kelas.price,
    )
    log_activity(request.user, 'updated', 'enrollment', new_enrollment.pk)
    messages.success(
        request,
        f'{student_name} berhasil dipindahkan dari {old_class_name} ke {target_kelas.name}.'
    )
    return redirect('admin_panel:enrollment_progress', enrollment_id=new_enrollment.pk)


# ── Announcements (admin manage) ──────────────────────────────────────────────

def _announcement_form_context(mode, ann=None, form_data=None):
    from announcements.models import Announcement
    return {
        'mode': mode,
        'ann': ann,            # may be None
        'form_data': form_data,  # may be None
        'target_role_choices': Announcement.TargetRole.choices,
        'level_choices': Announcement.TargetLevel.choices,
    }


def _parse_announcement_post(post):
    from announcements.models import Announcement
    title = post.get('title', '').strip()
    content = post.get('content', '').strip()
    target_role = post.get('target_role', Announcement.TargetRole.ALL)
    level = post.get('level', Announcement.TargetLevel.ALL)
    if target_role not in Announcement.TargetRole.values:
        target_role = Announcement.TargetRole.ALL
    if level not in Announcement.TargetLevel.values:
        level = Announcement.TargetLevel.ALL
    # Level only meaningful for STUDENT target
    if target_role != Announcement.TargetRole.STUDENT:
        level = Announcement.TargetLevel.ALL
    is_pinned = post.get('is_pinned') == '1'
    is_active = post.get('is_active') == '1'
    scheduled_at = post.get('scheduled_at', '').strip() or None
    expires_at = post.get('expires_at', '').strip() or None
    errors = []
    if not title:
        errors.append('Judul wajib diisi.')
    if not content:
        errors.append('Isi pengumuman wajib diisi.')
    return {
        'title': title, 'content': content, 'target_role': target_role,
        'level': level, 'is_pinned': is_pinned, 'is_active': is_active,
        'scheduled_at': scheduled_at, 'expires_at': expires_at,
    }, errors


@role_required('ADMIN')
def announcements_list(request):
    return render(request, 'admin_panel/announcements.html', {
        'q': request.GET.get('q', '').strip(),
        'role_filter': request.GET.get('role', ''),
        'status_filter': request.GET.get('status', ''),
    })


@role_required('ADMIN')
def announcements_list_partial(request):
    from announcements.models import Announcement
    q = request.GET.get('q', '').strip()
    role_filter = request.GET.get('role', '')
    status_filter = request.GET.get('status', '')

    qs = Announcement.objects.select_related('author').order_by('-is_pinned', '-created_at')
    if q:
        qs = qs.filter(Q(title__icontains=q) | Q(content__icontains=q))
    if role_filter in Announcement.TargetRole.values:
        qs = qs.filter(target_role=role_filter)
    if status_filter == 'active':
        qs = qs.filter(is_active=True)
    elif status_filter == 'inactive':
        qs = qs.filter(is_active=False)
    elif status_filter == 'pinned':
        qs = qs.filter(is_pinned=True)

    paginator = Paginator(qs, PAGE_SIZE)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    return render(request, 'admin_panel/_announcements_list_table.html', {
        'page_obj': page_obj,
        'q': q,
        'role_filter': role_filter,
        'status_filter': status_filter,
    })


@role_required('ADMIN')
def announcement_create(request):
    from announcements.models import Announcement
    if request.method == 'POST':
        data, errors = _parse_announcement_post(request.POST)
        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'admin_panel/announcement_form.html',
                          _announcement_form_context('create', form_data=request.POST))
        Announcement.objects.create(author=request.user, **data)
        log_activity(request.user, 'created', 'announcement')
        messages.success(request, f'Pengumuman "{data["title"]}" berhasil dipublikasikan.')
        return redirect('admin_panel:announcements_list')
    return render(request, 'admin_panel/announcement_form.html',
                  _announcement_form_context('create'))


@role_required('ADMIN')
def announcement_edit(request, pk):
    from announcements.models import Announcement
    ann = get_object_or_404(Announcement, pk=pk)
    if request.method == 'POST':
        data, errors = _parse_announcement_post(request.POST)
        if errors:
            for e in errors:
                messages.error(request, e)
            return render(request, 'admin_panel/announcement_form.html',
                          _announcement_form_context('edit', ann=ann, form_data=request.POST))
        for k, v in data.items():
            setattr(ann, k, v)
        ann.save()
        log_activity(request.user, 'updated', 'announcement', ann.pk)
        messages.success(request, 'Pengumuman berhasil diperbarui!')
        return redirect('admin_panel:announcements_list')
    return render(request, 'admin_panel/announcement_form.html',
                  _announcement_form_context('edit', ann=ann))


@role_required('ADMIN')
@require_POST
def announcement_delete(request, pk):
    from announcements.models import Announcement
    ann = get_object_or_404(Announcement, pk=pk)
    ann_pk = ann.pk
    ann.delete()
    log_activity(request.user, 'deleted', 'announcement', ann_pk)
    messages.success(request, 'Pengumuman berhasil dihapus.')
    return redirect('admin_panel:announcements_list')


@role_required('ADMIN')
@require_POST
def announcement_toggle(request, pk):
    """Toggle either `is_active` (default) or `is_pinned` per ?field=…"""
    from announcements.models import Announcement
    ann = get_object_or_404(Announcement, pk=pk)
    field = request.POST.get('field') or request.GET.get('field') or 'is_active'
    if field not in ('is_active', 'is_pinned'):
        field = 'is_active'
    setattr(ann, field, not getattr(ann, field))
    ann.save(update_fields=[field, 'updated_at'])
    log_activity(request.user, 'updated', 'announcement', ann.pk)
    return redirect('admin_panel:announcements_list')

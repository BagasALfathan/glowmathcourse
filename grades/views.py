import io
from collections import defaultdict

from django.contrib import messages
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from academics.models import Kelas
from accounts.decorators import role_required
from activity_logs.utils import log_activity
from enrollments.models import Enrollment, EnrollmentStatus
from sessions_app.models import Attendance, AttendanceStatus, Session

from .forms import GradeForm
from .models import Grade, GradeType


# ─── Teacher views ────────────────────────────────────────────────────────────

@role_required('TEACHER')
def teacher_grades_overview(request):
    """Quick-access list of all classes → click to manage grades per class."""
    from django.db.models import Count, Q
    from academics.models import KelasStatus
    qs = list(
        Kelas.objects
        .filter(teacher_profile__user=request.user, is_deleted=False)
        .select_related('subject')
        .annotate(
            active_enrolled=Count(
                'enrollments',
                filter=Q(
                    enrollments__status=EnrollmentStatus.ACTIVE,
                    enrollments__is_deleted=False,
                ),
                distinct=True,
            ),
            grade_count=Count('enrollments__grades', distinct=True),
        )
        .order_by('name')
    )
    active_klasses = [k for k in qs if k.status != KelasStatus.CLOSED]
    closed_klasses = [k for k in qs if k.status == KelasStatus.CLOSED]
    return render(request, 'grades/teacher_grades_overview.html', {
        'active_klasses': active_klasses,
        'closed_klasses': closed_klasses,
    })


@role_required('TEACHER')
def teacher_grades(request, pk):
    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)

    enrollments = (
        Enrollment.objects
        .filter(kelas=kelas, status=EnrollmentStatus.ACTIVE, is_deleted=False)
        .select_related('student_profile__user')
        .order_by('student_profile__user__last_name', 'student_profile__user__first_name')
    )

    # Prefetch grades per enrollment to avoid N+1
    enrollment_ids = [e.pk for e in enrollments]
    all_grades = (
        Grade.objects
        .filter(enrollment_id__in=enrollment_ids)
        .select_related('session')
        .order_by('grade_type', '-graded_at')
    )
    grades_by_enrollment = {}
    for grade in all_grades:
        grades_by_enrollment.setdefault(grade.enrollment_id, []).append(grade)

    rows = [
        {
            'enrollment': e,
            'grades': grades_by_enrollment.get(e.pk, []),
        }
        for e in enrollments
    ]

    return render(request, 'grades/teacher_grades.html', {
        'kelas': kelas,
        'rows': rows,
    })


@role_required('TEACHER')
def teacher_grade_create(request):
    # kelas_id comes from GET (initial load) or POST (hidden field on submit)
    kelas_id = request.POST.get('kelas_id') or request.GET.get('kelas_id')
    kelas = get_object_or_404(Kelas, pk=kelas_id, teacher_profile__user=request.user, is_deleted=False)

    form = GradeForm(request.POST or None, kelas=kelas)

    if request.method == 'POST' and form.is_valid():
        grade = form.save()
        log_activity(request.user, 'created', 'grade', grade.pk)
        messages.success(request, 'Nilai berhasil ditambahkan!')
        return redirect('grades:teacher_grades', pk=kelas.pk)

    return render(request, 'grades/teacher_grade_form.html', {
        'kelas': kelas,
        'form': form,
        'action': 'create',
        'form_title': 'Tambah Nilai',
    })


@role_required('TEACHER')
def teacher_grade_edit(request, pk):
    grade = get_object_or_404(Grade.objects.select_related('enrollment__kelas'), pk=pk)
    kelas = grade.enrollment.kelas

    if kelas.teacher != request.user:
        messages.error(request, 'Anda tidak memiliki akses untuk mengubah nilai ini.')
        return redirect('academics:teacher_classes')

    form = GradeForm(request.POST or None, instance=grade, kelas=kelas)

    if request.method == 'POST' and form.is_valid():
        form.save()
        log_activity(request.user, 'updated', 'grade', grade.pk)
        messages.success(request, 'Nilai berhasil diperbarui!')
        return redirect('grades:teacher_grades', pk=kelas.pk)

    return render(request, 'grades/teacher_grade_form.html', {
        'kelas': kelas,
        'form': form,
        'grade': grade,
        'action': 'edit',
        'form_title': 'Edit Nilai',
    })


@role_required('TEACHER')
@require_POST
def teacher_grade_delete(request, pk):
    grade = get_object_or_404(Grade.objects.select_related('enrollment__kelas'), pk=pk)
    kelas = grade.enrollment.kelas

    if kelas.teacher != request.user:
        messages.error(request, 'Anda tidak memiliki akses untuk menghapus nilai ini.')
        return redirect('academics:teacher_classes')

    grade_pk = grade.pk
    grade.delete()
    log_activity(request.user, 'deleted', 'grade', grade_pk)
    messages.success(request, 'Nilai berhasil dihapus.')
    return redirect('grades:teacher_grades', pk=kelas.pk)


# ─── Student views ─────────────────────────────────────────────────────────────

@role_required('STUDENT')
def my_grades(request):
    all_enrollments = list(
        Enrollment.objects
        .filter(
            student_profile__user=request.user,
            status__in=[EnrollmentStatus.ACTIVE, EnrollmentStatus.COMPLETED],
            is_deleted=False,
        )
        .select_related('kelas__subject', 'kelas__teacher_profile__user')
        .order_by('kelas__name')
    )
    enrollment_ids = [e.pk for e in all_enrollments]

    all_grades = list(
        Grade.objects
        .filter(enrollment_id__in=enrollment_ids)
        .select_related('session')
        .order_by('grade_type', '-graded_at')
    )
    grades_by_enrollment = {}
    for grade in all_grades:
        grades_by_enrollment.setdefault(grade.enrollment_id, []).append(grade)

    def _build_rows(enrollments):
        rows = []
        for e in enrollments:
            grades = grades_by_enrollment.get(e.pk, [])
            avg = round(sum(float(g.score) for g in grades) / len(grades), 1) if grades else None
            rows.append({'enrollment': e, 'grades': grades, 'avg': avg})
        return rows

    active_rows = _build_rows([e for e in all_enrollments if e.status == EnrollmentStatus.ACTIVE])
    completed_rows = _build_rows([e for e in all_enrollments if e.status == EnrollmentStatus.COMPLETED])

    recent_grades = (
        Grade.objects
        .filter(enrollment_id__in=enrollment_ids)
        .select_related('enrollment__kelas__subject')
        .order_by('-graded_at')[:5]
    )

    return render(request, 'grades/my_grades.html', {
        'active_rows': active_rows,
        'completed_rows': completed_rows,
        'recent_grades': recent_grades,
    })


@role_required('STUDENT')
def my_grades_detail(request, kelas_id):
    kelas = get_object_or_404(Kelas, pk=kelas_id, is_deleted=False)
    enrollment = get_object_or_404(
        Enrollment,
        student_profile__user=request.user,
        kelas=kelas,
        is_deleted=False,
    )

    grades = list(
        Grade.objects
        .filter(enrollment=enrollment)
        .select_related('session')
        .order_by('grade_type', '-graded_at')
    )

    scores = [float(g.score) for g in grades]
    avg = round(sum(scores) / len(scores), 1) if scores else None
    highest = max(scores, default=None)
    lowest = min(scores, default=None)

    return render(request, 'grades/my_grades_detail.html', {
        'kelas': kelas,
        'enrollment': enrollment,
        'grades': grades,
        'avg': avg,
        'highest': highest,
        'lowest': lowest,
        'total': len(grades),
    })


# ─── HTMX inline grade edit ────────────────────────────────────────────────────

@role_required('TEACHER')
def teacher_grade_inline_edit(request, pk):
    """Return inline score edit form (GET), or display cell if cancel=true."""
    grade = get_object_or_404(Grade.objects.select_related('enrollment__kelas'), pk=pk)
    if grade.enrollment.kelas.teacher != request.user:
        return HttpResponse(status=403)
    if request.GET.get('cancel') == 'true':
        return render(request, 'grades/_grade_score_display.html', {'grade': grade})
    return render(request, 'grades/_grade_score_edit.html', {'grade': grade})


@role_required('TEACHER')
@require_POST
def teacher_grade_inline_save(request, pk):
    """Save inline score edit and return the updated display cell (HTMX POST)."""
    grade = get_object_or_404(Grade.objects.select_related('enrollment__kelas'), pk=pk)
    if grade.enrollment.kelas.teacher != request.user:
        return HttpResponse(status=403)

    score_raw = request.POST.get('score', '').strip()
    try:
        score = float(score_raw)
        if not (0 <= score <= 100):
            raise ValueError
    except ValueError:
        return render(request, 'grades/_grade_score_edit.html', {
            'grade': grade,
            'error': 'Nilai harus antara 0 dan 100.',
        })

    grade.score = score
    grade.save(update_fields=['score', 'updated_at'])
    log_activity(request.user, 'updated', 'grade', grade.pk)
    return render(request, 'grades/_grade_score_display.html', {'grade': grade})


# ─── Grades export helpers ────────────────────────────────────────────────────

_TYPE_ORDER = [GradeType.QUIZ, GradeType.MIDTERM, GradeType.FINAL, GradeType.ASSIGNMENT]
_TYPE_LABELS = {
    GradeType.QUIZ: 'Kuis',
    GradeType.MIDTERM: 'UTS',
    GradeType.FINAL: 'UAS',
    GradeType.ASSIGNMENT: 'Tugas',
}


def _build_grades_columns_and_rows(kelas, enrollments):
    """
    Returns (columns, rows) where:
    - columns: list of (grade_type, index, label)
    - rows: list of {'name': str, 'cells': [...], 'avg': float|None}
    """
    enrollment_ids = [e.pk for e in enrollments]
    all_grades = list(
        Grade.objects
        .filter(enrollment_id__in=enrollment_ids)
        .order_by('grade_type', 'graded_at')
    )

    # group by (enrollment_id, grade_type)
    by_key = defaultdict(list)
    for g in all_grades:
        by_key[(g.enrollment_id, g.grade_type)].append(g)

    # max count per type across all enrollments
    type_max = defaultdict(int)
    for (_, gt), gs in by_key.items():
        type_max[gt] = max(type_max[gt], len(gs))

    # build column list in fixed order
    columns = []
    for gt in _TYPE_ORDER:
        count = type_max.get(gt, 0)
        label_base = _TYPE_LABELS[gt]
        for i in range(count):
            label = f'{label_base} {i + 1}' if count > 1 else label_base
            columns.append((gt, i, label))

    rows = []
    for enrollment in enrollments:
        cells = []
        scores = []
        for gt, idx, _ in columns:
            gs = by_key.get((enrollment.pk, gt), [])
            if idx < len(gs):
                val = float(gs[idx].score)
                cells.append(val)
                scores.append(val)
            else:
                cells.append(None)
        avg = round(sum(scores) / len(scores), 1) if scores else None
        rows.append({
            'name': enrollment.student.get_full_name(),
            'cells': cells,
            'avg': avg,
        })

    return columns, rows


@role_required('TEACHER')
def export_grades_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)
    enrollments = list(
        Enrollment.objects
        .filter(kelas=kelas, status=EnrollmentStatus.ACTIVE, is_deleted=False)
        .select_related('student_profile__user')
        .order_by('student_profile__user__last_name', 'student_profile__user__first_name')
    )
    columns, rows = _build_grades_columns_and_rows(kelas, enrollments)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Nilai'

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center')

    headers = ['No', 'Nama Siswa'] + [col[2] for col in columns] + ['Rata-rata']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    for i, row in enumerate(rows, 1):
        cells = [v if v is not None else '-' for v in row['cells']]
        avg = row['avg'] if row['avg'] is not None else '-'
        ws.append([i, row['name']] + cells + [avg])

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    for col_idx in range(3, 3 + len(columns) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = kelas.name.replace(' ', '_').replace('/', '-')
    date_str = timezone.localdate().strftime('%Y%m%d')
    filename = f'Nilai_{safe_name}_{date_str}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@role_required('TEACHER')
def export_grades_pdf(request, pk):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)
    enrollments = list(
        Enrollment.objects
        .filter(kelas=kelas, status=EnrollmentStatus.ACTIVE, is_deleted=False)
        .select_related('student_profile__user')
        .order_by('student_profile__user__last_name', 'student_profile__user__first_name')
    )
    columns, rows = _build_grades_columns_and_rows(kelas, enrollments)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()

    def para(text, style='Normal', **kwargs):
        s = ParagraphStyle('_', parent=styles[style], **kwargs)
        return Paragraph(text, s)

    teacher_name = kelas.teacher.get_full_name() if kelas.teacher else '-'
    period_name = kelas.academic_period.name if kelas.academic_period_id else '-'

    elements = [
        para(f'Laporan Nilai — {kelas.name}', 'Heading1', fontSize=14, alignment=1),
        Spacer(1, 0.2 * cm),
        para(
            f'{kelas.subject.name}  ·  Guru: {teacher_name}  ·  Periode: {period_name}',
            fontSize=9, alignment=1,
        ),
        Spacer(1, 0.5 * cm),
    ]

    header = ['No', 'Nama Siswa'] + [col[2] for col in columns] + ['Rata-rata']
    data = [header]
    for i, row in enumerate(rows, 1):
        cells = [f'{v:.1f}' if v is not None else '-' for v in row['cells']]
        avg = f"{row['avg']}" if row['avg'] is not None else '-'
        data.append([str(i), row['name']] + cells + [avg])

    indent_col = colors.HexColor('#4F46E5')
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), indent_col),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(para(
        f'Dicetak: {timezone.localdate().strftime("%d %B %Y")}',
        fontSize=7, alignment=2,
    ))

    doc.build(elements)
    buffer.seek(0)

    safe_name = kelas.name.replace(' ', '_').replace('/', '-')
    date_str = timezone.localdate().strftime('%Y%m%d')
    filename = f'Nilai_{safe_name}_{date_str}.pdf'
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Student print views ──────────────────────────────────────────────────────

@role_required('STUDENT')
def print_my_grades(request):
    enrollments = list(
        Enrollment.objects
        .filter(student_profile__user=request.user, is_deleted=False)
        .exclude(status=EnrollmentStatus.DROPPED)
        .select_related('kelas__subject', 'kelas__teacher_profile__user', 'kelas__academic_period')
        .order_by('kelas__name')
    )
    enrollment_ids = [e.pk for e in enrollments]
    all_grades = list(
        Grade.objects
        .filter(enrollment_id__in=enrollment_ids)
        .order_by('grade_type', 'graded_at')
    )
    grades_by_enrollment = defaultdict(list)
    for g in all_grades:
        grades_by_enrollment[g.enrollment_id].append(g)

    rows = []
    for e in enrollments:
        grades = grades_by_enrollment.get(e.pk, [])
        avg = round(sum(float(g.score) for g in grades) / len(grades), 1) if grades else None
        rows.append({'enrollment': e, 'grades': grades, 'avg': avg})

    return render(request, 'grades/print_my_grades.html', {
        'rows': rows,
        'student': request.user,
        'printed_date': timezone.localdate(),
    })


# ─── Progress report ──────────────────────────────────────────────────────────

_PROGRESS_TYPE_ORDER = [GradeType.QUIZ, GradeType.MIDTERM, GradeType.FINAL, GradeType.ASSIGNMENT]
_PROGRESS_TYPE_LABELS = {
    GradeType.QUIZ:       'Kuis',
    GradeType.MIDTERM:    'UTS',
    GradeType.FINAL:      'UAS',
    GradeType.ASSIGNMENT: 'Tugas',
}


def _build_progress_ctx(enrollment):
    """Build all data needed for the student progress report."""
    kelas = enrollment.kelas

    # ── Attendance ────────────────────────────────────────────────────────────
    sessions = list(Session.objects.filter(kelas=kelas).order_by('session_number'))
    attendance_map = {
        a.session_id: a
        for a in Attendance.objects.filter(enrollment=enrollment).select_related('session')
    }

    present_count   = sum(1 for s in sessions if attendance_map.get(s.pk) and attendance_map[s.pk].status == AttendanceStatus.PRESENT)
    permitted_count = sum(1 for s in sessions if attendance_map.get(s.pk) and attendance_map[s.pk].status == AttendanceStatus.PERMITTED)
    absent_count    = sum(1 for s in sessions if attendance_map.get(s.pk) and attendance_map[s.pk].status == AttendanceStatus.ABSENT)
    unmarked_count  = sum(1 for s in sessions if s.pk not in attendance_map)
    total_sessions  = len(sessions)

    attendance_pct = round((present_count / total_sessions) * 100) if total_sessions > 0 else 0
    attendance_good = attendance_pct >= 75

    session_rows = [
        {
            'session': s,
            'attendance': attendance_map.get(s.pk),
            'status': attendance_map[s.pk].status if s.pk in attendance_map else None,
            'status_label': attendance_map[s.pk].get_status_display() if s.pk in attendance_map else 'Belum',
        }
        for s in sessions
    ]

    # ── Grades ────────────────────────────────────────────────────────────────
    grades_qs = list(Grade.objects.filter(enrollment=enrollment).order_by('grade_type', 'graded_at'))
    by_type = {}
    for g in grades_qs:
        by_type.setdefault(g.grade_type, []).append(g)

    all_scores = []
    grade_type_rows = []
    for gt in _PROGRESS_TYPE_ORDER:
        gs = by_type.get(gt, [])
        if gs:
            scores = [float(g.score) for g in gs]
            avg = round(sum(scores) / len(scores), 1)
            all_scores.extend(scores)
            grade_type_rows.append({'type': gt, 'label': _PROGRESS_TYPE_LABELS[gt], 'grades': gs, 'avg': avg})

    overall_avg = round(sum(all_scores) / len(all_scores), 1) if all_scores else None

    if overall_avg is None:
        letter = '-'
    elif overall_avg >= 85:
        letter = 'A'
    elif overall_avg >= 70:
        letter = 'B'
    elif overall_avg >= 55:
        letter = 'C'
    else:
        letter = 'D'

    if overall_avg is not None and total_sessions > 0:
        if overall_avg >= 75 and attendance_good:
            summary = f'Siswa menunjukkan performa baik dengan rata-rata nilai {overall_avg} dan kehadiran {attendance_pct}%.'
        elif overall_avg < 75 and attendance_good:
            summary = f'Kehadiran baik ({attendance_pct}%), namun rata-rata nilai perlu ditingkatkan ({overall_avg}).'
        elif overall_avg >= 75 and not attendance_good:
            summary = f'Rata-rata nilai baik ({overall_avg}), namun kehadiran perlu ditingkatkan ({attendance_pct}%).'
        else:
            summary = f'Perlu perhatian lebih — rata-rata nilai {overall_avg} dengan kehadiran {attendance_pct}%.'
    elif overall_avg is not None:
        summary = f'Rata-rata nilai: {overall_avg}. Belum ada data pertemuan.'
    elif total_sessions > 0:
        summary = f'Kehadiran {attendance_pct}%. Belum ada nilai yang diinput.'
    else:
        summary = 'Belum ada data nilai maupun kehadiran.'

    return {
        'enrollment': enrollment,
        'kelas': kelas,
        'student': enrollment.student,
        'session_rows': session_rows,
        'present_count': present_count,
        'permitted_count': permitted_count,
        'absent_count': absent_count,
        'unmarked_count': unmarked_count,
        'total_sessions': total_sessions,
        'attendance_pct': attendance_pct,
        'attendance_good': attendance_good,
        'grade_type_rows': grade_type_rows,
        'overall_avg': overall_avg,
        'letter': letter,
        'summary': summary,
    }


def _progress_enrollment_qs():
    return Enrollment.objects.select_related(
        'student_profile__user', 'kelas__subject', 'kelas__teacher_profile__user', 'kelas__academic_period'
    )


@role_required('TEACHER')
def teacher_student_progress(request, pk, enrollment_id):
    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)
    enrollment = get_object_or_404(_progress_enrollment_qs(), pk=enrollment_id, kelas=kelas)
    ctx = _build_progress_ctx(enrollment)
    ctx['back_url'] = f'/teacher/classes/{kelas.pk}/students/'
    ctx['print_url'] = f'/teacher/classes/{kelas.pk}/students/{enrollment_id}/progress/print/'
    ctx['pdf_url'] = f'/teacher/classes/{kelas.pk}/students/{enrollment_id}/progress/pdf/'
    return render(request, 'grades/teacher_student_progress.html', ctx)


@role_required('TEACHER')
def teacher_student_progress_print(request, pk, enrollment_id):
    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)
    enrollment = get_object_or_404(_progress_enrollment_qs(), pk=enrollment_id, kelas=kelas)
    ctx = _build_progress_ctx(enrollment)
    ctx['printed_date'] = timezone.localdate()
    return render(request, 'grades/progress_print.html', ctx)


@role_required('TEACHER')
def teacher_student_progress_pdf(request, pk, enrollment_id):
    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)
    enrollment = get_object_or_404(_progress_enrollment_qs(), pk=enrollment_id, kelas=kelas)
    return _generate_progress_pdf(enrollment)


@role_required('STUDENT')
def student_progress(request, kelas_id):
    kelas = get_object_or_404(Kelas, pk=kelas_id, is_deleted=False)
    enrollment = get_object_or_404(
        _progress_enrollment_qs(), student_profile__user=request.user, kelas=kelas,
    )
    ctx = _build_progress_ctx(enrollment)
    ctx['back_url'] = '/my-classes/'
    ctx['print_url'] = f'/my-progress/{kelas_id}/print/'
    return render(request, 'grades/student_progress.html', ctx)


@role_required('STUDENT')
def student_progress_print(request, kelas_id):
    kelas = get_object_or_404(Kelas, pk=kelas_id, is_deleted=False)
    enrollment = get_object_or_404(
        _progress_enrollment_qs(), student_profile__user=request.user, kelas=kelas,
    )
    ctx = _build_progress_ctx(enrollment)
    ctx['printed_date'] = timezone.localdate()
    return render(request, 'grades/progress_print.html', ctx)


def _generate_progress_pdf(enrollment):
    """Build and return an HTTP PDF response for the progress report."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    ctx = _build_progress_ctx(enrollment)
    kelas   = enrollment.kelas
    student = enrollment.student

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
    )
    styles = getSampleStyleSheet()
    indigo = colors.HexColor('#4F46E5')
    light  = colors.HexColor('#F9FAFB')
    gray   = colors.HexColor('#6B7280')
    rule   = colors.HexColor('#E5E7EB')

    def p(text, size=9, bold=False, color=colors.black, align=0):
        s = ParagraphStyle(
            '_', parent=styles['Normal'],
            fontSize=size,
            fontName='Helvetica-Bold' if bold else 'Helvetica',
            textColor=color, alignment=align,
            spaceAfter=0,
        )
        return Paragraph(text, s)

    teacher_name = kelas.teacher.get_full_name() if kelas.teacher else '—'
    period_name  = kelas.academic_period.name if kelas.academic_period_id else '—'
    profile      = getattr(student, 'student_profile', None)

    elems = [
        p('Laporan Progress Siswa', size=16, bold=True, align=1),
        Spacer(1, 0.25 * cm),
        p(kelas.name, size=12, bold=True, align=1),
        p(f'{kelas.subject.name}  ·  Guru: {teacher_name}  ·  Periode: {period_name}', size=9, color=gray, align=1),
        Spacer(1, 0.3 * cm),
        HRFlowable(width='100%', thickness=0.5, color=rule),
        Spacer(1, 0.4 * cm),
        p('Informasi Siswa', size=11, bold=True),
        Spacer(1, 0.15 * cm),
    ]

    student_rows = [
        ['Nama',   student.get_full_name()],
        ['Email',  student.email],
        ['Jenjang', profile.get_level_display() if profile else '—'],
        ['Sekolah', profile.school_name if profile and profile.school_name else '—'],
        ['Status',  enrollment.get_status_display()],
    ]
    t_info = Table(student_rows, colWidths=[4 * cm, 12 * cm])
    t_info.setStyle(TableStyle([
        ('FONTSIZE',     (0, 0), (-1, -1), 9),
        ('TEXTCOLOR',    (0, 0), (0, -1),  gray),
        ('TOPPADDING',   (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 2),
    ]))
    elems += [t_info, Spacer(1, 0.5 * cm),
              HRFlowable(width='100%', thickness=0.3, color=rule),
              Spacer(1, 0.3 * cm),
              p('Kehadiran', size=11, bold=True),
              Spacer(1, 0.15 * cm)]

    att_head = [['Total Pertemuan', 'Hadir', 'Izin', 'Alpha', 'Belum', 'Kehadiran %']]
    att_data = [[
        str(ctx['total_sessions']),
        str(ctx['present_count']),
        str(ctx['permitted_count']),
        str(ctx['absent_count']),
        str(ctx['unmarked_count']),
        f"{ctx['attendance_pct']}%",
    ]]
    t_att = Table(att_head + att_data, colWidths=[3 * cm] * 6)
    t_att.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), indigo),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, -1), 9),
        ('ALIGN',        (0, 0), (-1, -1), 'CENTER'),
        ('GRID',         (0, 0), (-1, -1), 0.3, rule),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
    ]))
    elems += [t_att, Spacer(1, 0.25 * cm)]

    if ctx['session_rows']:
        sess_data = [['Pertemuan', 'Topik', 'Status']]
        for row in ctx['session_rows']:
            s = row['session']
            sess_data.append([f"Pertemuan ke-{s.session_number}", s.topic or '—', row['status_label']])
        t_sess = Table(sess_data, colWidths=[4 * cm, 10 * cm, 4 * cm])
        t_sess.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#F3F4F6')),
            ('FONTNAME',      (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, light]),
            ('GRID',          (0, 0), (-1, -1), 0.3, rule),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        elems.append(t_sess)

    elems += [Spacer(1, 0.5 * cm),
              HRFlowable(width='100%', thickness=0.3, color=rule),
              Spacer(1, 0.3 * cm),
              p('Nilai', size=11, bold=True),
              Spacer(1, 0.15 * cm)]

    if ctx['grade_type_rows']:
        for type_row in ctx['grade_type_rows']:
            elems.append(p(f"— {type_row['label']}", size=9, bold=True))
            gd = [['#', 'Nilai', 'Catatan']]
            for i, g in enumerate(type_row['grades'], 1):
                gd.append([str(i), f"{float(g.score):.1f}", g.notes or '—'])
            gd.append(['', f"Rata-rata: {type_row['avg']}", ''])
            t_g = Table(gd, colWidths=[1.5 * cm, 4 * cm, 12.5 * cm])
            t_g.setStyle(TableStyle([
                ('BACKGROUND',    (0, 0),  (-1, 0),  colors.HexColor('#F3F4F6')),
                ('FONTNAME',      (0, 0),  (-1, 0),  'Helvetica-Bold'),
                ('FONTNAME',      (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE',      (0, 0),  (-1, -1), 8),
                ('ALIGN',         (1, 0),  (1, -1),  'CENTER'),
                ('GRID',          (0, 0),  (-1, -2), 0.3, rule),
                ('TOPPADDING',    (0, 0),  (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0),  (-1, -1), 3),
            ]))
            elems += [t_g, Spacer(1, 0.2 * cm)]
    else:
        elems.append(p('Belum ada nilai yang diinput.', color=gray))

    avg_str = str(ctx['overall_avg']) if ctx['overall_avg'] is not None else '—'
    overall_data = [
        ['Rata-rata Keseluruhan', 'Grade', 'Status Kehadiran'],
        [avg_str, ctx['letter'], 'Baik' if ctx['attendance_good'] else 'Kurang'],
    ]
    t_overall = Table(overall_data, colWidths=[6 * cm, 4 * cm, 8 * cm])
    t_overall.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, 0), indigo),
        ('TEXTCOLOR',     (0, 0), (-1, 0), colors.white),
        ('FONTNAME',      (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1, -1), 10),
        ('ALIGN',         (0, 0), (-1, -1), 'CENTER'),
        ('GRID',          (0, 0), (-1, -1), 0.3, rule),
        ('TOPPADDING',    (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elems += [
        Spacer(1, 0.3 * cm), t_overall,
        Spacer(1, 0.3 * cm), p(ctx['summary'], color=gray),
        Spacer(1, 0.5 * cm),
        HRFlowable(width='100%', thickness=0.3, color=rule),
        Spacer(1, 0.2 * cm),
        p(f'Dicetak: {timezone.localdate().strftime("%d %B %Y")}', size=7, color=gray, align=2),
    ]

    doc.build(elems)
    buffer.seek(0)

    safe_student = student.get_full_name().replace(' ', '_')
    safe_kelas   = kelas.name.replace(' ', '_').replace('/', '-')
    date_str     = timezone.localdate().strftime('%Y%m%d')
    filename     = f'Progress_{safe_student}_{safe_kelas}_{date_str}.pdf'
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

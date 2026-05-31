import datetime
import io
import json

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import models, transaction
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST

from academics.models import Kelas
from accounts.decorators import role_required
from activity_logs.utils import log_activity
from enrollments.models import Enrollment, EnrollmentStatus

from .forms import SessionForm
from .models import (
    Attendance, AttendanceStatus, BookingKind, BookingStatus,
    Session, SessionBooking, SessionStatus, SessionType,
)

_WEEKDAY_TO_DAY = ['MONDAY', 'TUESDAY', 'WEDNESDAY', 'THURSDAY', 'FRIDAY', 'SATURDAY', 'SUNDAY']


def _session_overlap_conflicts(teacher_profile, date, start_time, end_time, exclude_session_id=None):
    """Return Session rows on `date` (across all teacher's classes) that overlap
    with [start_time, end_time). Strict `<` comparison — back-to-back times
    (one session ends exactly when next begins) are NOT considered overlap.
    """
    qs = (
        Session.objects
        .filter(
            kelas__teacher_profile=teacher_profile,
            kelas__is_deleted=False,
            date=date,
        )
        .select_related('kelas')
    )
    if exclude_session_id:
        qs = qs.exclude(pk=exclude_session_id)
    conflicts = []
    for s in qs:
        if s.start_time is None or s.end_time is None:
            continue  # legacy / malformed — skip rather than crash
        # Overlap iff start_a < end_b AND start_b < end_a
        if start_time < s.end_time and s.start_time < end_time:
            conflicts.append(s)
    return conflicts


@role_required('TEACHER')
def teacher_attendance_overview(request):
    """Quick-access list of all classes → click to go to sessions/attendance."""
    from django.db.models import Count
    from academics.models import KelasStatus

    today = timezone.localdate()

    klasses = list(
        Kelas.objects
        .filter(teacher_profile__user=request.user, is_deleted=False)
        .select_related('subject')
        .order_by('name')
    )
    kelas_ids = [k.pk for k in klasses]

    sessions_raw = list(
        Session.objects
        .filter(kelas_id__in=kelas_ids)
        .annotate(att_count=Count('attendances'))
        .order_by('kelas_id', '-date')
    )
    sessions_by_kelas = {}
    for s in sessions_raw:
        sessions_by_kelas.setdefault(s.kelas_id, []).append(s)

    active_rows = []
    closed_rows = []
    for kelas in klasses:
        sessions = sessions_by_kelas.get(kelas.pk, [])
        latest_session = sessions[0] if sessions else None
        incomplete_count = sum(
            1 for s in sessions
            if s.date <= today
            and s.status == SessionStatus.COMPLETED
            and s.att_count == 0
        )
        row = {
            'kelas': kelas,
            'latest_session': latest_session,
            'total_sessions': len(sessions),
            'incomplete_count': incomplete_count,
        }
        if kelas.status == KelasStatus.CLOSED:
            closed_rows.append(row)
        else:
            active_rows.append(row)

    return render(request, 'sessions_app/teacher_attendance_overview.html', {
        'active_rows': active_rows,
        'closed_rows': closed_rows,
        'today': today,
    })


@role_required('TEACHER')
def teacher_sessions(request, pk):
    """Phase 3B — Manage Sessions page.

    Single page: progress bar + existing-sessions preview + repeat form
    to bulk-add new sessions. Strict cap: existing + new ≤ kelas.total_sessions.
    Race-safe via select_for_update inside transaction.

    URL stays `sessions_app:teacher_sessions` (5 existing templates reference
    this name — renaming would break links). Spec called for `academics:`
    namespace but pragmatic to keep here.
    """
    teacher_profile = request.user.teacher_profile
    kelas = get_object_or_404(
        Kelas.objects.select_related('subject', 'academic_period'),
        pk=pk, teacher_profile=teacher_profile, is_deleted=False,
    )

    existing_sessions = Session.objects.filter(kelas=kelas).order_by('session_number')
    existing_count = existing_sessions.count()
    remaining = max(kelas.total_sessions - existing_count, 0)

    valid_types = {v for v, _ in SessionType.choices}
    valid_statuses = {v for v, _ in SessionStatus.choices}

    if request.method == 'POST':
        rows = []
        i = 0
        while f'topic_{i}' in request.POST:
            rows.append({
                'topic': (request.POST.get(f'topic_{i}') or '').strip(),
                'date': request.POST.get(f'date_{i}') or '',
                'start_time': request.POST.get(f'start_time_{i}') or '',
                'end_time': request.POST.get(f'end_time_{i}') or '',
                'session_type': request.POST.get(f'session_type_{i}') or SessionType.REGULAR,
                'status': request.POST.get(f'status_{i}') or SessionStatus.SCHEDULED,
            })
            i += 1

        # Drop completely-empty rows (user clicked + Tambah Baris but didn't fill)
        rows = [r for r in rows if r['topic'] or r['date']]

        errors = []
        if not rows:
            errors.append('Tambah minimal 1 baris sesi.')
        if existing_count + len(rows) > kelas.total_sessions:
            errors.append(
                f'Total sesi melebihi kapasitas kelas '
                f'({existing_count + len(rows)} > {kelas.total_sessions}). '
                f'Hapus beberapa baris atau ubah Total Pertemuan di Edit Kelas.'
            )
        batch_intervals = []  # (date, start, end) accumulator for intra-batch overlap check
        for idx, r in enumerate(rows):
            n = idx + 1
            if not r['topic']:
                errors.append(f'Sesi baris #{n}: Topik wajib diisi.')
            if not r['date']:
                errors.append(f'Sesi baris #{n}: Tanggal wajib diisi.')
            if not r['start_time'] or not r['end_time']:
                errors.append(f'Sesi baris #{n}: Jam mulai & selesai wajib.')
            if r['session_type'] not in valid_types:
                errors.append(f'Sesi baris #{n}: Tipe sesi tidak valid.')
            if r['status'] not in valid_statuses:
                errors.append(f'Sesi baris #{n}: Status tidak valid.')

            # Stop here if basic field errors are already present for this row —
            # parsing/overlap checks would be noise.
            if not (r['date'] and r['start_time'] and r['end_time']):
                continue

            # Parse to typed values. HTML5 inputs post `YYYY-MM-DD` and `HH:MM`,
            # but some browsers also include seconds (`HH:MM:SS`). Try both.
            try:
                r_date = datetime.datetime.strptime(r['date'], '%Y-%m-%d').date()
            except (TypeError, ValueError):
                errors.append(f'Sesi baris #{n}: format tanggal tidak valid.')
                continue
            try:
                r_start = datetime.datetime.strptime(r['start_time'][:5], '%H:%M').time()
                r_end = datetime.datetime.strptime(r['end_time'][:5], '%H:%M').time()
            except (TypeError, ValueError):
                errors.append(f'Sesi baris #{n}: format jam tidak valid.')
                continue

            # Replace string values with typed objects so bulk_create gets clean inputs.
            r['date'] = r_date
            r['start_time'] = r_start
            r['end_time'] = r_end

            # End must be strictly after start
            if r_end <= r_start:
                errors.append(f'Sesi baris #{n}: Jam selesai harus setelah jam mulai.')
                continue

            # (1) Cross-class overlap vs existing sessions in the DB
            conflicts = _session_overlap_conflicts(
                teacher_profile, r_date, r_start, r_end,
            )
            if conflicts:
                c = conflicts[0]
                errors.append(
                    f'Sesi baris #{n} ({r_date} {r_start.strftime("%H:%M")}–{r_end.strftime("%H:%M")}) '
                    f'bentrok dengan "{c.kelas.name}" sesi #{c.session_number} '
                    f'({c.start_time.strftime("%H:%M")}–{c.end_time.strftime("%H:%M")}) di tanggal yang sama.'
                )
                continue

            # (2) Intra-batch overlap vs rows already accepted in this submit
            intra_clash = False
            for b_idx, (b_date, b_start, b_end) in enumerate(batch_intervals):
                if r_date == b_date and r_start < b_end and b_start < r_end:
                    errors.append(
                        f'Sesi baris #{n} bentrok dengan baris #{b_idx + 1} di batch ini '
                        f'(tanggal {r_date}, jam tumpang tindih).'
                    )
                    intra_clash = True
                    break
            if not intra_clash:
                batch_intervals.append((r_date, r_start, r_end))

        if errors:
            for e in errors:
                messages.error(request, e)
        else:
            try:
                with transaction.atomic():
                    kelas_locked = Kelas.objects.select_for_update().get(pk=kelas.pk)
                    current_count = Session.objects.filter(kelas=kelas_locked).count()
                    # Defense in depth — recheck cap inside lock
                    if current_count + len(rows) > kelas_locked.total_sessions:
                        raise ValueError(
                            f'Kapasitas terlampaui ({current_count + len(rows)} > '
                            f'{kelas_locked.total_sessions}). Ada yang menambahkan sesi '
                            f'lebih dulu — refresh halaman.'
                        )
                    new_sessions = [
                        Session(
                            kelas=kelas_locked,
                            session_number=current_count + offset + 1,
                            topic=r['topic'],
                            date=r['date'],
                            start_time=r['start_time'],
                            end_time=r['end_time'],
                            session_type=r['session_type'],
                            status=r['status'],
                            capacity=kelas_locked.capacity,
                        )
                        for offset, r in enumerate(rows)
                    ]
                    Session.objects.bulk_create(new_sessions)
                log_activity(request.user, 'created', 'session', kelas.pk)
                messages.success(request, f'✓ {len(rows)} sesi berhasil ditambahkan!')
                return redirect('sessions_app:teacher_sessions', pk=kelas.pk)
            except Exception as e:
                messages.error(request, f'Gagal menyimpan: {e}')

        # On error: re-render with submitted rows preserved
        return render(request, 'sessions_app/teacher_sessions.html', {
            'kelas': kelas,
            'existing_sessions': existing_sessions,
            'existing_count': existing_count,
            'remaining': remaining,
            'progress_pct': int(existing_count * 100 / kelas.total_sessions) if kelas.total_sessions else 0,
            'submitted_rows': rows,
            'session_types': SessionType.choices,
            'statuses': SessionStatus.choices,
        })

    return render(request, 'sessions_app/teacher_sessions.html', {
        'kelas': kelas,
        'existing_sessions': existing_sessions,
        'existing_count': existing_count,
        'remaining': remaining,
        'progress_pct': int(existing_count * 100 / kelas.total_sessions) if kelas.total_sessions else 0,
        'submitted_rows': [],
        'session_types': SessionType.choices,
        'statuses': SessionStatus.choices,
    })


@role_required('TEACHER')
def teacher_session_row_partial(request, pk):
    """HTMX endpoint: returns one fresh form-row HTML for `+ Tambah Baris`.

    Enforces the session cap server-side. If existing + already-drafted rows
    has already filled the kelas's total_sessions, returns empty body + an
    HX-Trigger header carrying a warning message for the client to flash.
    """
    teacher_profile = request.user.teacher_profile
    kelas = get_object_or_404(
        Kelas, pk=pk, teacher_profile=teacher_profile, is_deleted=False,
    )

    try:
        idx = max(int(request.GET.get('idx', 0)), 0)
    except (TypeError, ValueError):
        idx = 0

    existing = Session.objects.filter(kelas=kelas).count()
    # `idx` is the count of rows ALREADY in the DOM (button sends current count
    # BEFORE appending). So `existing + idx` is the post-add session count if
    # we were to render. If that would exceed cap, refuse + flash warning.
    if existing + idx >= kelas.total_sessions:
        resp = HttpResponse('')
        resp['HX-Trigger'] = json.dumps({
            'showCapWarning': f'Kapasitas {kelas.total_sessions} sesi sudah tercapai.',
        })
        return resp

    return render(request, 'sessions_app/_session_row_partial.html', {
        'idx': idx,
        'next_num': existing + idx + 1,
        'session_types': SessionType.choices,
        'statuses': SessionStatus.choices,
        'row': {},  # no prefill
    })


@role_required('TEACHER')
def teacher_session_create(request, kelas_id):
    kelas = get_object_or_404(Kelas, pk=kelas_id, teacher_profile__user=request.user, is_deleted=False)

    # Determine next session number
    last_session = Session.objects.filter(kelas=kelas).order_by('-session_number').first()
    next_number = (last_session.session_number + 1) if last_session else 1

    # Guard: cannot exceed total_sessions
    if next_number > kelas.total_sessions:
        messages.warning(request, 'Semua pertemuan sudah dibuat.')
        return redirect('sessions_app:teacher_sessions', pk=kelas.pk)

    form = SessionForm(request.POST or None, kelas=kelas)

    if request.method == 'POST' and form.is_valid():
        session_date = form.cleaned_data['date']
        if session_date < kelas.start_date or session_date > kelas.end_date:
            form.add_error(
                'date',
                f'Tanggal harus antara {kelas.start_date.strftime("%d %b %Y")} '
                f'dan {kelas.end_date.strftime("%d %b %Y")}.',
            )
        else:
            session = form.save(commit=False)
            session.kelas = kelas
            session.session_number = next_number
            session.status = SessionStatus.SCHEDULED
            session.save()
            messages.success(request, f'Pertemuan ke-{next_number} berhasil dibuat!')
            return redirect('sessions_app:teacher_sessions', pk=kelas.pk)

    return render(request, 'sessions_app/teacher_session_create.html', {
        'kelas': kelas,
        'form': form,
        'next_number': next_number,
    })


@role_required('TEACHER')
def teacher_session_edit(request, pk):
    session = get_object_or_404(Session.objects.select_related('kelas'), pk=pk)
    if session.kelas.teacher != request.user:
        messages.error(request, 'Anda tidak memiliki akses untuk pertemuan ini.')
        return redirect('academics:teacher_classes')

    kelas = session.kelas

    form = SessionForm(request.POST or None, instance=session, kelas=kelas)

    if request.method == 'POST' and form.is_valid():
        session_date = form.cleaned_data['date']
        if session_date < kelas.start_date or session_date > kelas.end_date:
            form.add_error(
                'date',
                f'Tanggal harus antara {kelas.start_date.strftime("%d %b %Y")} '
                f'dan {kelas.end_date.strftime("%d %b %Y")}.',
            )
        else:
            form.save()
            messages.success(request, f'Pertemuan ke-{session.session_number} berhasil diperbarui!')
            return redirect('sessions_app:teacher_sessions', pk=kelas.pk)

    return render(request, 'sessions_app/teacher_session_edit.html', {
        'kelas': kelas,
        'session': session,
        'form': form,
    })


@role_required('TEACHER')
@require_POST
def teacher_session_update_status(request, pk):
    session = get_object_or_404(Session, pk=pk)

    # Ownership check: session's kelas must belong to this teacher
    if session.kelas.teacher != request.user:
        messages.error(request, 'Anda tidak memiliki akses untuk mengubah pertemuan ini.')
        return redirect('academics:teacher_classes')

    new_status = request.POST.get('status', '').strip()
    if new_status not in SessionStatus.values:
        messages.error(request, 'Status tidak valid.')
        return redirect('sessions_app:teacher_sessions', pk=session.kelas_id)

    if session.status == new_status:
        return redirect('sessions_app:teacher_sessions', pk=session.kelas_id)

    session.status = new_status
    session.save(update_fields=['status', 'updated_at'])

    if new_status == SessionStatus.CANCELLED:
        cancelled_count = SessionBooking.objects.filter(
            session=session, status=BookingStatus.BOOKED
        ).update(status=BookingStatus.CANCELLED)
        status_label = dict(SessionStatus.choices).get(new_status, new_status)
        messages.success(
            request,
            f'Pertemuan ke-{session.session_number} dibatalkan. '
            f'{cancelled_count} pemesanan siswa ikut dibatalkan.'
        )
    else:
        status_label = dict(SessionStatus.choices).get(new_status, new_status)
        messages.success(
            request,
            f'Pertemuan ke-{session.session_number} diubah menjadi {status_label}.'
        )
    return redirect('sessions_app:teacher_sessions', pk=session.kelas_id)


@role_required('TEACHER')
def teacher_attendance(request, pk):
    """Phase 3B — Per-session attendance marking.

    Single POST with `status_<enrollment_id>` fields, 3-state toggle per
    student (PRESENT/PERMITTED/ABSENT). All active enrollments are listed
    (DROPPED + soft-deleted excluded). Future sessions emit a warning but
    still allow pre-marking (locked decision).
    """
    teacher_profile = request.user.teacher_profile
    session = get_object_or_404(
        Session.objects.select_related('kelas'),
        pk=pk,
        kelas__teacher_profile=teacher_profile,
        kelas__is_deleted=False,
    )
    kelas = session.kelas

    enrollments = list(
        Enrollment.objects
        .filter(kelas=kelas, is_deleted=False)
        .exclude(status=EnrollmentStatus.DROPPED)
        .select_related('student_profile__user')
        .order_by('student_profile__user__first_name', 'student_profile__user__last_name')
    )

    valid_statuses = {v for v, _ in AttendanceStatus.choices}
    today = timezone.localdate()
    is_future = session.date > today

    if request.method == 'POST':
        try:
            with transaction.atomic():
                for enr in enrollments:
                    status = (request.POST.get(f'status_{enr.id}') or '').strip()
                    if status not in valid_statuses:
                        # Empty / invalid → skip (allows partial marking)
                        continue
                    Attendance.objects.update_or_create(
                        enrollment=enr,
                        session=session,
                        defaults={
                            'status': status,
                            'marked_by': request.user,
                        },
                    )
            log_activity(request.user, 'marked_attendance', 'session', session.pk)
            messages.success(
                request,
                f'✓ Kehadiran sesi #{session.session_number} tersimpan!',
            )
            return redirect('sessions_app:teacher_attendance', pk=session.pk)
        except Exception as e:
            messages.error(request, f'Gagal menyimpan: {e}')

    # Existing attendance → prefill the toggle per enrollment
    existing = {
        a.enrollment_id: a.status
        for a in Attendance.objects.filter(session=session)
    }
    marked_count = 0
    for enr in enrollments:
        enr.current_status = existing.get(enr.id, '')
        if enr.current_status:
            marked_count += 1

    return render(request, 'sessions_app/teacher_attendance.html', {
        'session': session,
        'kelas': kelas,
        'enrollments': enrollments,
        'statuses': AttendanceStatus.choices,
        'marked_count': marked_count,
        'total_students': len(enrollments),
        'is_future': is_future,
    })


# ─── Role-aware Session Detail (named student_session_redirect for backward URL compat) ───

@login_required
def student_session_redirect(request, pk):
    """Role-aware Session Detail page.

    - Student enrolled in this kelas: shows their attendance status for this session.
    - Teacher of this kelas: shows all attendances + link to edit absensi.
    - Admin: read-only view of all attendances.
    - Student NOT enrolled: redirect to class_detail with an error message.
    """
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    from django.utils import timezone
    from django.db.models import Q
    from academics.models import KelasStatus
    from accounts.models import Role
    from enrollments.models import Enrollment, EnrollmentStatus

    session = get_object_or_404(
        Session.objects.select_related('kelas__subject', 'kelas__teacher_profile__user'),
        pk=pk,
    )
    kelas = session.kelas
    user = request.user
    today = timezone.localdate()

    is_teacher = (
        user.role == Role.TEACHER
        and kelas.teacher_profile.user_id == user.pk
    )
    is_admin = (user.role == Role.ADMIN)
    is_student = (user.role == Role.STUDENT)

    enrollment = None
    my_attendance = None
    all_attendances = []
    present_count = permitted_count = absent_count = 0

    if is_student:
        enrollment = (
            Enrollment.objects
            .filter(
                student_profile__user=user,
                kelas=kelas,
                status=EnrollmentStatus.ACTIVE,
                is_deleted=False,
            )
            .first()
        )
        if enrollment is None:
            messages.error(request, 'Kamu tidak terdaftar di kelas ini.')
            return redirect('academics:class_detail', pk=kelas.pk)
        my_attendance = (
            Attendance.objects
            .filter(session=session, enrollment=enrollment)
            .first()
        )
    elif is_teacher or is_admin:
        all_attendances = list(
            Attendance.objects
            .filter(session=session)
            .select_related('enrollment__student_profile__user')
            .order_by('enrollment__student_profile__user__first_name')
        )
        for a in all_attendances:
            if a.status == AttendanceStatus.PRESENT:
                present_count += 1
            elif a.status == AttendanceStatus.PERMITTED:
                permitted_count += 1
            elif a.status == AttendanceStatus.ABSENT:
                absent_count += 1
    else:
        return redirect('dashboard:router')

    # Session ordering: this session's position + prev/next
    all_session_ids = list(
        Session.objects
        .filter(kelas=kelas)
        .order_by('date', 'start_time')
        .values_list('pk', flat=True)
    )
    try:
        idx = all_session_ids.index(session.pk)
    except ValueError:
        idx = 0
    session_number = idx + 1
    total_sessions = len(all_session_ids)
    prev_session = (
        Session.objects.select_related('kelas').get(pk=all_session_ids[idx - 1])
        if idx > 0 else None
    )
    next_session = (
        Session.objects.select_related('kelas').get(pk=all_session_ids[idx + 1])
        if idx < total_sessions - 1 else None
    )

    return render(request, 'sessions_app/session_detail.html', {
        'session': session,
        'kelas': kelas,
        'today': today,
        'session_number': session_number,
        'total_sessions': total_sessions,
        'is_teacher': is_teacher,
        'is_admin': is_admin,
        'is_student': is_student,
        'enrollment': enrollment,
        'my_attendance': my_attendance,
        'all_attendances': all_attendances,
        'present_count': present_count,
        'permitted_count': permitted_count,
        'absent_count': absent_count,
        'prev_session': prev_session,
        'next_session': next_session,
    })


@role_required('STUDENT')
def student_session_list(request, enrollment_id):
    from enrollments.models import Enrollment, EnrollmentStatus
    enrollment = get_object_or_404(
        Enrollment,
        pk=enrollment_id,
        student_profile__user=request.user,
        status=EnrollmentStatus.ACTIVE,
        is_deleted=False,
    )
    kelas = enrollment.kelas

    sessions = list(
        Session.objects
        .filter(kelas=kelas)
        .annotate(
            booked_count_ann=Count('bookings', filter=models.Q(bookings__status=BookingStatus.BOOKED), distinct=True),
        )
        .order_by('session_number')
    )

    # Build schedule lookup by day name
    schedules_by_day = {s.day: s for s in kelas.schedules.all()}

    # Map session_id → booking for this student
    my_bookings = {
        b.session_id: b
        for b in SessionBooking.objects.filter(enrollment=enrollment)
    }

    today = timezone.localdate()
    current_time = timezone.localtime().time()
    rows = []
    for session in sessions:
        booking = my_bookings.get(session.pk)
        is_booked = booking and booking.status == BookingStatus.BOOKED
        is_full = session.capacity > 0 and session.booked_count_ann >= session.capacity and not is_booked
        day_name = _WEEKDAY_TO_DAY[session.date.weekday()]
        schedule = schedules_by_day.get(day_name)
        st = session.start_time or (schedule.start_time if schedule else None)
        is_past_start = (
            session.is_today
            and st is not None
            and current_time >= st
        )
        rows.append({
            'session': session,
            'booking': booking,
            'is_booked': is_booked,
            'is_full': is_full,
            'schedule': schedule,
            'is_past_start': is_past_start,
        })

    total_booked = sum(1 for r in rows if r['is_booked'])

    # Countdown for the next booked upcoming session
    now_aware = timezone.localtime()
    next_countdown_start_ts = None
    next_countdown_end_ts = None
    for row in rows:
        if row['is_booked'] and not row['session'].is_past:
            s = row['session']
            sched = row['schedule']
            # Prefer session's own stored times; fall back to class schedule
            st = s.start_time or (sched.start_time if sched else None)
            et = s.end_time or (sched.end_time if sched else None)
            if not (st and et):
                continue
            aware_start = timezone.make_aware(datetime.datetime.combine(s.date, st))
            aware_end = timezone.make_aware(datetime.datetime.combine(s.date, et))
            if aware_end > now_aware:
                next_countdown_start_ts = int(aware_start.timestamp() * 1000)
                next_countdown_end_ts = int(aware_end.timestamp() * 1000)
                break

    return render(request, 'sessions_app/student_session_list.html', {
        'kelas': kelas,
        'enrollment': enrollment,
        'rows': rows,
        'total_booked': total_booked,
        'today': today,
        'next_countdown_start_ts': next_countdown_start_ts,
        'next_countdown_end_ts': next_countdown_end_ts,
    })


@role_required('STUDENT')
@require_POST
def student_book_session(request, enrollment_id, session_id):
    from enrollments.models import Enrollment, EnrollmentStatus
    enrollment = get_object_or_404(
        Enrollment,
        pk=enrollment_id,
        student_profile__user=request.user,
        status=EnrollmentStatus.ACTIVE,
        is_deleted=False,
    )
    session = get_object_or_404(Session, pk=session_id, kelas=enrollment.kelas)
    today = timezone.localdate()

    # Validate: cannot book past sessions
    if session.date < today:
        messages.error(request, 'Tidak bisa mendaftar pertemuan yang sudah lewat.')
        return redirect('sessions_app:student_session_list', enrollment_id=enrollment_id)

    # Validate: cannot book today's session if it has already started
    if session.date == today:
        schedules_by_day = {s.day: s for s in enrollment.kelas.schedules.all()}
        day_name = _WEEKDAY_TO_DAY[session.date.weekday()]
        schedule = schedules_by_day.get(day_name)
        st = session.start_time or (schedule.start_time if schedule else None)
        if st and timezone.localtime().time() >= st:
            messages.error(request, 'Pertemuan sudah dimulai, tidak bisa mendaftar.')
            return redirect('sessions_app:student_session_list', enrollment_id=enrollment_id)

    # Validate: cannot book cancelled sessions
    if session.status == SessionStatus.CANCELLED:
        messages.error(request, 'Pertemuan ini sudah dibatalkan.')
        return redirect('sessions_app:student_session_list', enrollment_id=enrollment_id)

    # Pre-check capacity before attempting create
    if session.capacity > 0:
        already_booked = SessionBooking.objects.filter(
            session=session, status=BookingStatus.BOOKED
        ).count()
        # Check if student already has an active booking (exclude from count)
        student_has_booking = SessionBooking.objects.filter(
            enrollment=enrollment, session=session, status=BookingStatus.BOOKED
        ).exists()
        if not student_has_booking and already_booked >= session.capacity:
            messages.error(request, 'Maaf, pertemuan ini sudah penuh.')
            return redirect('sessions_app:student_session_list', enrollment_id=enrollment_id)

    booking, created = SessionBooking.objects.get_or_create(
        enrollment=enrollment,
        session=session,
        defaults={'status': BookingStatus.BOOKED, 'kind': BookingKind.PICKED},
    )

    if not created:
        if booking.status == BookingStatus.CANCELLED:
            booking.status = BookingStatus.BOOKED
            # Reactivating a cancelled booking — treat as a deliberate re-pick.
            booking.kind = BookingKind.PICKED
            booking.save(update_fields=['status', 'kind', 'updated_at'])
            messages.success(request, f'Berhasil mendaftar ke Pertemuan ke-{session.session_number}!')
        else:
            messages.info(request, 'Kamu sudah terdaftar di pertemuan ini.')
    else:
        # Race-condition guard: re-check after insert
        booked = SessionBooking.objects.filter(session=session, status=BookingStatus.BOOKED).count()
        if session.capacity > 0 and booked > session.capacity:
            booking.status = BookingStatus.CANCELLED
            booking.save(update_fields=['status', 'updated_at'])
            messages.error(request, 'Maaf, pertemuan ini sudah penuh.')
        else:
            messages.success(request, f'Berhasil mendaftar ke Pertemuan ke-{session.session_number}!')

    return redirect('sessions_app:student_session_list', enrollment_id=enrollment_id)


@role_required('STUDENT')
@require_POST
def student_cancel_booking(request, enrollment_id, session_id):
    from enrollments.models import Enrollment, EnrollmentStatus
    enrollment = get_object_or_404(
        Enrollment,
        pk=enrollment_id,
        student_profile__user=request.user,
        status=EnrollmentStatus.ACTIVE,
        is_deleted=False,
    )
    session = get_object_or_404(Session, pk=session_id, kelas=enrollment.kelas)
    today = timezone.localdate()

    # Cannot cancel today's or past sessions
    if session.date <= today:
        messages.error(request, 'Tidak bisa membatalkan pertemuan yang sudah berlangsung atau hari ini.')
        return redirect('sessions_app:student_session_list', enrollment_id=enrollment_id)

    booking = SessionBooking.objects.filter(enrollment=enrollment, session=session).first()
    if booking and booking.status == BookingStatus.BOOKED:
        booking.status = BookingStatus.CANCELLED
        booking.save(update_fields=['status', 'updated_at'])
        messages.success(request, f'Pendaftaran Pertemuan ke-{session.session_number} dibatalkan.')
    else:
        messages.info(request, 'Tidak ada pendaftaran aktif untuk pertemuan ini.')

    return redirect('sessions_app:student_session_list', enrollment_id=enrollment_id)


# ─── Phase 3R: helpers + session-first pick ────────────────────────────────────

def _auto_book_regular_sessions(enrollment):
    """For every REGULAR session in enrollment's kelas, ensure a
    SessionBooking(kind=AUTO, status=BOOKED) exists. Capacity governance is at
    the Kelas level (already enforced by enroll()), so we do NOT block per
    session here. Uses bulk_create(ignore_conflicts=True) so the unique
    constraint (enrollment, session) makes the call idempotent — re-enroll
    after drop, or partial pre-existing PICKED rows, are tolerated.

    Returns the number of newly-created bookings (best-effort; bulk_create
    with ignore_conflicts returns None on SQLite for the inserted rows, so
    we count via a pre/post diff).
    """
    regular_sessions = Session.objects.filter(
        kelas=enrollment.kelas,
        session_type=SessionType.REGULAR,
        status__in=[SessionStatus.SCHEDULED, SessionStatus.COMPLETED],
    )
    if not regular_sessions.exists():
        return 0
    pre = SessionBooking.objects.filter(enrollment=enrollment).count()
    rows = [
        SessionBooking(
            enrollment=enrollment, session=s,
            status=BookingStatus.BOOKED, kind=BookingKind.AUTO,
        )
        for s in regular_sessions
    ]
    SessionBooking.objects.bulk_create(rows, ignore_conflicts=True)
    post = SessionBooking.objects.filter(enrollment=enrollment).count()
    return max(0, post - pre)


@role_required('STUDENT')
@require_POST
def student_pick_session(request, session_id):
    """Session-first enrollment: student picks ONE Session → we ensure a
    class-level Enrollment exists (creating if needed, race-safe via the
    shared _try_enroll helper) → mark the picked booking kind=PICKED → fan
    out AUTO bookings for the kelas's other REGULAR sessions.

    Mirrors enroll() pre-checks (level match, kelas OPEN, not started,
    schedule conflict) so behavior is consistent regardless of which entry
    point the student uses.
    """
    from enrollments.views import _student_schedule_conflict, _try_enroll
    from academics.models import KelasStatus

    student_profile = request.user.student_profile
    session = get_object_or_404(
        Session.objects.select_related('kelas'),
        pk=session_id,
        kelas__is_deleted=False,
        status=SessionStatus.SCHEDULED,
    )
    kelas = session.kelas

    # Mirror enroll()'s pre-checks
    if kelas.status != KelasStatus.OPEN:
        if kelas.status == KelasStatus.FULL:
            messages.error(request, 'Kelas sudah penuh.')
        else:
            messages.error(request, 'Kelas ini sudah tidak menerima pendaftaran.')
        return redirect('academics:class_detail', pk=kelas.pk)

    if kelas.start_date < timezone.localdate():
        messages.error(request, 'Pendaftaran sudah ditutup, kelas sudah dimulai.')
        return redirect('academics:class_detail', pk=kelas.pk)

    if student_profile.level != kelas.level:
        messages.error(
            request,
            f'Kelas ini untuk jenjang {kelas.level}, bukan {student_profile.level}.',
        )
        return redirect('academics:class_detail', pk=kelas.pk)

    # Schedule-conflict check (skipped if student is already enrolled — let
    # _try_enroll handle the 'already' path without re-flagging conflicts).
    existing = Enrollment.objects.filter(
        student_profile=student_profile, kelas=kelas, is_deleted=False
    ).first()
    if not existing or existing.status != EnrollmentStatus.ACTIVE:
        conflict = _student_schedule_conflict(student_profile, kelas)
        if conflict:
            messages.error(request, conflict)
            return redirect('academics:class_detail', pk=kelas.pk)

    # Race-safe Enrollment ensure (Kelas + Enrollment row both locked inside).
    result, payload = _try_enroll(student_profile, kelas)

    if result == 'full':
        messages.error(request, 'Kelas sudah penuh.')
        return redirect('academics:class_detail', pk=kelas.pk)
    if result == 'closed':
        messages.error(request, 'Kelas ini sudah tidak menerima pendaftaran.')
        return redirect('academics:class_detail', pk=kelas.pk)
    if result == 'completed':
        messages.error(request, 'Anda sudah menyelesaikan kelas ini.')
        return redirect('academics:class_detail', pk=kelas.pk)
    # 'already' and 'ok' both yield a valid enrollment to attach the booking to
    enrollment = payload

    # Session-level capacity check (slot-aware for the session-first flow)
    if session.capacity and session.capacity > 0:
        already_at_session = (
            SessionBooking.objects
            .filter(session=session, status=BookingStatus.BOOKED, is_deleted=False)
            .exclude(enrollment=enrollment)
            .count()
        )
        if already_at_session >= session.capacity:
            messages.error(request, 'Sesi ini sudah penuh.')
            return redirect('academics:class_detail', pk=kelas.pk)

    with transaction.atomic():
        # The picked session: PICKED (or upgrade an existing AUTO/CANCELLED → PICKED+BOOKED)
        booking, created = SessionBooking.objects.get_or_create(
            enrollment=enrollment,
            session=session,
            defaults={'status': BookingStatus.BOOKED, 'kind': BookingKind.PICKED},
        )
        if not created:
            # Pre-existing row (AUTO from prior class-enroll, or CANCELLED) — promote it.
            dirty = []
            if booking.status != BookingStatus.BOOKED:
                booking.status = BookingStatus.BOOKED
                dirty.append('status')
            if booking.kind != BookingKind.PICKED:
                booking.kind = BookingKind.PICKED
                dirty.append('kind')
            if booking.is_deleted:
                booking.is_deleted = False
                booking.deleted_at = None
                dirty += ['is_deleted', 'deleted_at']
            if dirty:
                dirty.append('updated_at')
                booking.save(update_fields=dirty)

        # Fan out AUTO bookings for the rest of the kelas's REGULAR sessions.
        # bulk_create(ignore_conflicts=True) honors unique (enrollment, session) —
        # the picked row above stays PICKED because get_or_create won't update it.
        _auto_book_regular_sessions(enrollment)

    log_activity(request.user, 'created', 'session_booking', booking.pk)
    if result == 'ok':
        messages.success(
            request,
            f'Berhasil bergabung di kelas {kelas.name} — Pertemuan ke-{session.session_number} terdaftar.',
        )
    else:
        messages.success(
            request,
            f'Pertemuan ke-{session.session_number} berhasil didaftarkan.',
        )
    return redirect('enrollments:my_class_detail', enrollment_id=enrollment.pk)


# ─── Attendance exports ────────────────────────────────────────────────────────

def _build_attendance_data(kelas):
    """Return (sessions, enrollments, att_map) for export views."""
    sessions = list(Session.objects.filter(kelas=kelas).order_by('session_number'))
    enrollments = list(
        Enrollment.objects
        .filter(kelas=kelas, status=EnrollmentStatus.ACTIVE, is_deleted=False)
        .select_related('student_profile__user')
        .order_by('student_profile__user__last_name', 'student_profile__user__first_name')
    )
    att_map = {
        (a.enrollment_id, a.session_id): a.status
        for a in Attendance.objects.filter(session__kelas=kelas)
    }
    return sessions, enrollments, att_map


def _attendance_row(enrollment, sessions, att_map):
    """Return (cells_list, present, permitted, absent) for one student."""
    cells = []
    present = permitted = absent = 0
    for session in sessions:
        status = att_map.get((enrollment.pk, session.pk))
        if status == AttendanceStatus.PRESENT:
            cells.append('H')
            present += 1
        elif status == AttendanceStatus.PERMITTED:
            cells.append('I')
            permitted += 1
        elif status == AttendanceStatus.ABSENT:
            cells.append('A')
            absent += 1
        else:
            cells.append('-')
    return cells, present, permitted, absent


@role_required('TEACHER')
def export_attendance_excel(request, pk):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)
    sessions, enrollments, att_map = _build_attendance_data(kelas)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kehadiran'

    header_fill = PatternFill('solid', fgColor='4F46E5')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    center = Alignment(horizontal='center', vertical='center')

    # Header
    headers = ['No', 'Nama Siswa']
    for s in sessions:
        headers.append(f'P{s.session_number}')
    headers += ['Total H', 'Total I', 'Total A', '% Hadir']
    ws.append(headers)

    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Data rows
    for i, enrollment in enumerate(enrollments, 1):
        cells, present, permitted, absent = _attendance_row(enrollment, sessions, att_map)
        total_marked = present + permitted + absent
        pct = f'{round((present + permitted) / total_marked * 100)}%' if total_marked else '-'
        row = [i, enrollment.student.get_full_name()] + cells + [present, permitted, absent, pct]
        ws.append(row)

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    for col_idx in range(3, 3 + len(sessions)):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 5
    for offset in range(4):
        col_letter = ws.cell(row=1, column=3 + len(sessions) + offset).column_letter
        ws.column_dimensions[col_letter].width = 10

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = kelas.name.replace(' ', '_').replace('/', '-')
    date_str = timezone.localdate().strftime('%Y%m%d')
    filename = f'Kehadiran_{safe_name}_{date_str}.xlsx'
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@role_required('TEACHER')
def export_attendance_pdf(request, pk):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    kelas = get_object_or_404(Kelas, pk=pk, teacher_profile__user=request.user, is_deleted=False)
    sessions, enrollments, att_map = _build_attendance_data(kelas)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()

    def para(text, style='Normal', **kwargs):
        s = ParagraphStyle('_', parent=styles[style], **kwargs)
        return Paragraph(text, s)

    teacher_name = kelas.teacher.get_full_name() if kelas.teacher else '-'
    period_name = kelas.academic_period.name if kelas.academic_period_id else '-'

    elements = [
        para(f'Laporan Kehadiran — {kelas.name}', 'Heading1', fontSize=14, alignment=1),
        Spacer(1, 0.2 * cm),
        para(
            f'{kelas.subject.name}  ·  Guru: {teacher_name}  ·  Periode: {period_name}',
            fontSize=9, alignment=1,
        ),
        Spacer(1, 0.5 * cm),
    ]

    header = ['No', 'Nama Siswa'] + [f'P{s.session_number}' for s in sessions] + ['H', 'I', 'A', '%']
    data = [header]

    for i, enrollment in enumerate(enrollments, 1):
        cells, present, permitted, absent = _attendance_row(enrollment, sessions, att_map)
        total_marked = present + permitted + absent
        pct = f'{round((present + permitted) / total_marked * 100)}%' if total_marked else '-'
        data.append([str(i), enrollment.student.get_full_name()] + cells + [str(present), str(permitted), str(absent), pct])

    indent_col = colors.HexColor('#4F46E5')
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), indent_col),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(para(
        f'Dicetak: {timezone.localdate().strftime("%d %B %Y")}',
        fontSize=7, alignment=2,
    ))

    doc.build(elements)
    buffer.seek(0)

    safe_name = kelas.name.replace(' ', '_').replace('/', '-')
    date_str = timezone.localdate().strftime('%Y%m%d')
    filename = f'Kehadiran_{safe_name}_{date_str}.pdf'
    response = HttpResponse(buffer.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
